

# app.py
# Full file: keeps your original app logic intact, and adds:
#  - Road model (name, speed limit, center lat/lon, radius)
#  - OverspeedEvent model (records events per road)
#  - Floating minimizable UI widget in DASHBOARD_HTML (search/create/select road, view all)
#  - Admin endpoints: create/list/delete/search roads
#  - Automatic overspeed recording in heartbeat (per-road, deduped by snapshot)
#  - Report endpoints: download full-app Excel/PDF, or per-road Excel/PDF
#
# IMPORTANT: This file is meant to replace your current app.py exactly.
# All original code is preserved; new code is added where noted.
#
# Start of file -------------------------------------------------------------

import os
import uuid
import threading
from math import radians, sin, cos, atan2, sqrt, degrees
from datetime import datetime, timedelta
import json
import time
from collections import defaultdict
import io

from flask import (
    Flask, request, jsonify, render_template_string, abort,
    redirect, url_for, session, flash, send_file, Response
)
from flask_sqlalchemy import SQLAlchemy
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.security import generate_password_hash, check_password_hash
try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

# New imports for reports & excel
try:
    import pandas as pd
except Exception:
    pd = None

# reportlab optional for PDF generation
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# -------------------------
# Configuration (tunable)
# -------------------------
DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///beacon.db")
CLEANUP_STALE_SECONDS = int(os.environ.get("CLEANUP_STALE_SECONDS", "12"))  # remove snapshots older than this
NEARBY_DEFAULT_RADIUS_M = float(os.environ.get("NEARBY_DEFAULT_RADIUS_M", "1000"))  # 1 km default per your spec
HEARTBEAT_MIN_INTERVAL_S = float(os.environ.get("HEARTBEAT_MIN_INTERVAL_S", "0.5"))  # basic rate-limit
UNSAFE_TTC_SECONDS = float(os.environ.get("UNSAFE_TTC_SECONDS", "6.0"))  # threshold for opposite-direction unsafe
CONFIRMATION_RADIUS_M = float(os.environ.get("CONFIRMATION_RADIUS_M", "30.0"))  # support devices gathering radius
# Optional API token for scripted admin calls
ADMIN_API_TOKEN = os.environ.get("ADMIN_API_TOKEN")
# Optionally seed admin via env on first run:
ADMIN_USER = os.environ.get("ADMIN_USER")
ADMIN_PASS = os.environ.get("ADMIN_PASS")

VEHICLE_LENGTH_M = float(os.environ.get("VEHICLE_LENGTH_M", "5.0"))
OVERTAKE_EXTRA_M = float(os.environ.get("OVERTAKE_EXTRA_M", "5.0"))
SAFETY_FACTOR = float(os.environ.get("SAFETY_FACTOR", "1.5"))

# Accident detection tuning (server-side inference; no native changes required)
# thresholds in m/s^2 and m/s
ACCIDENT_DECEL_HIGH_MPS2 = float(os.environ.get("ACCIDENT_DECEL_HIGH_MPS2", "8.0"))   # strong deceleration ~ -8 m/s^2
ACCIDENT_DECEL_MED_MPS2 = float(os.environ.get("ACCIDENT_DECEL_MED_MPS2", "5.0"))     # medium ~ -5 m/s^2
ACCIDENT_SPEED_DROP_MPS = float(os.environ.get("ACCIDENT_SPEED_DROP_MPS", "10.0"))    # drop of 10 m/s (~36 km/h)
ACCIDENT_BEARING_JUMP_DEG = float(os.environ.get("ACCIDENT_BEARING_JUMP_DEG", "60.0"))# abrupt heading change
ACCIDENT_TIME_WINDOW_S = float(os.environ.get("ACCIDENT_TIME_WINDOW_S", "3.0"))       # examine last N seconds

# Jam detection tuning (new)
JAM_DETECT_INTERVAL_S = float(os.environ.get("JAM_DETECT_INTERVAL_S", "5.0"))
JAM_MIN_DEVICES = int(os.environ.get("JAM_MIN_DEVICES", "3"))
JAM_SPEED_THRESHOLD_MPS = float(os.environ.get("JAM_SPEED_THRESHOLD_MPS", "3.0"))  # < ~11 km/h
JAM_CLUSTER_RADIUS_M = float(os.environ.get("JAM_CLUSTER_RADIUS_M", "50.0"))
JAM_RETENTION_S = int(os.environ.get("JAM_RETENTION_S", "60"))  # keep jams recent for this many seconds

# -------------------------
# Flask + SQLAlchemy + SocketIO init
# -------------------------
app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.secret_key = os.environ.get("FLASK_SECRET", "dev-secret-change-me")

# Use threading async_mode for Windows/dev. For production (multi-process) configure SocketIO with Redis and eventlet.
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")

db = SQLAlchemy(app)

# -------------------------
# Models
# -------------------------
class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(128), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Device(db.Model):
    id = db.Column(db.String(36), primary_key=True)  # UUID4 hex
    token = db.Column(db.String(64), unique=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    owner = db.Column(db.String(128))
    car_name = db.Column(db.String(128))
    car_model = db.Column(db.String(128))
    plate = db.Column(db.String(64))
    extra = db.Column(db.Text)
    revoked = db.Column(db.Boolean, default=False)

class Snapshot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    device_id = db.Column(db.String(36), db.ForeignKey('device.id'), index=True)
    ts = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    lat = db.Column(db.Float)
    lon = db.Column(db.Float)
    speed_mps = db.Column(db.Float)   # store m/s
    bearing_deg = db.Column(db.Float) # 0..360
    heading_deg = db.Column(db.Float) # optional
    source = db.Column(db.String(32), default="app") # e.g., "app", "web"
    raw = db.Column(db.Text) # JSON dump of raw payload (optional)

# -------------------------
# New models: Road + OverspeedEvent
# -------------------------
class Road(db.Model):
    id = db.Column(db.String(36), primary_key=True, default=lambda: uuid.uuid4().hex)
    name = db.Column(db.String(256), nullable=False, index=True)
    speed_limit_kmh = db.Column(db.Float, nullable=False)  # stored in km/h for user-friendliness
    # Simple circular area for road monitoring (center + radius)
    center_lat = db.Column(db.Float, nullable=True)
    center_lon = db.Column(db.Float, nullable=True)
    radius_m = db.Column(db.Float, nullable=True, default=50.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class OverspeedEvent(db.Model):
    id = db.Column(db.String(36), primary_key=True, default=lambda: uuid.uuid4().hex)
    device_id = db.Column(db.String(36), db.ForeignKey('device.id'))
    road_id = db.Column(db.String(36), db.ForeignKey('road.id'))
    snapshot_id = db.Column(db.Integer, db.ForeignKey('snapshot.id'), nullable=True)
    ts = db.Column(db.DateTime, default=datetime.utcnow)
    speed_kmh = db.Column(db.Float)
    lat = db.Column(db.Float)
    lon = db.Column(db.Float)
    raw = db.Column(db.Text)


# -------------------------
# Additional auth / traffic models
# -------------------------
class PoliceUser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(128), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class GKUser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(128), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class TrafficZone(db.Model):
    id = db.Column(db.String(36), primary_key=True, default=lambda: uuid.uuid4().hex)
    name = db.Column(db.String(256), nullable=False, index=True)
    scope = db.Column(db.String(64), nullable=False, default="custom")
    center_lat = db.Column(db.Float, nullable=True)
    center_lon = db.Column(db.Float, nullable=True)
    radius_m = db.Column(db.Float, nullable=True)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class BootstrapState(db.Model):
    key = db.Column(db.String(64), primary_key=True)
    value = db.Column(db.String(64), nullable=False, default="1")
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class BroadcastMessage(db.Model):
    id = db.Column(db.String(36), primary_key=True, default=lambda: uuid.uuid4().hex)
    title = db.Column(db.String(255), nullable=False)
    body = db.Column(db.Text, nullable=False)
    target_type = db.Column(db.String(32), nullable=False)  # single/all/overspeeders/zone/road/county/search
    target_value = db.Column(db.String(255))
    creator_role = db.Column(db.String(32))
    creator_username = db.Column(db.String(128))
    recipient_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class BroadcastDelivery(db.Model):
    id = db.Column(db.String(36), primary_key=True, default=lambda: uuid.uuid4().hex)
    message_id = db.Column(db.String(36), db.ForeignKey('broadcast_message.id'), index=True)
    device_id = db.Column(db.String(36), db.ForeignKey('device.id'), index=True)
    delivered_at = db.Column(db.DateTime, nullable=True)
    read_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# -------------------------
# Bootstrap state helpers
# -------------------------
_BOOTSTRAP_STATE_KEY = 'bootstrap_open'


def _bootstrap_state_row(create_if_missing=False):
    """Return the bootstrap state row, creating it when requested."""
    row = BootstrapState.query.filter_by(key=_BOOTSTRAP_STATE_KEY).first()
    if row is None and create_if_missing:
        row = BootstrapState(key=_BOOTSTRAP_STATE_KEY, value='1')
        db.session.add(row)
        db.session.commit()
    return row


def _close_bootstrap():
    """Mark bootstrap registration as closed, but never fail startup if the DB is not ready."""
    try:
        row = _bootstrap_state_row(create_if_missing=True)
        if row is not None:
            row.value = '0'
            db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
# -------------------------
# Simple in-memory rate tracking (per-device)
# -------------------------
_last_heartbeat_at = {}  # device_id -> timestamp (float)
# Note: for multi-process deployments, move this into Redis or DB.

# -------------------------
# In-memory active device cache (thread-safe)
# -------------------------
active_devices = {}
active_devices_lock = threading.Lock()

def update_active_device_from_snapshot(snap):
    """Given a Snapshot model instance, update in-memory active_devices."""
    try:
        with active_devices_lock:
            active_devices[snap.device_id] = {
                "ts": snap.ts,
                "lat": snap.lat,
                "lon": snap.lon,
                "speed_mps": snap.speed_mps,
                "bearing_deg": snap.bearing_deg,
                "source": snap.source,
                "raw": (json.loads(snap.raw) if snap.raw else None)
            }
    except Exception:
        pass

def prune_active_devices():
    """Remove entries older than CLEANUP_STALE_SECONDS."""
    try:
        cutoff = datetime.utcnow() - timedelta(seconds=CLEANUP_STALE_SECONDS)
        with active_devices_lock:
            to_del = [k for k,v in active_devices.items() if v.get("ts") < cutoff]
            for k in to_del:
                active_devices.pop(k, None)
    except Exception:
        pass

# -------------------------
# In-memory alert & jam stores (new)
# -------------------------
alerts_store = []  # list of alert dicts (accident alerts)
alerts_lock = threading.Lock()

jams_store = []    # list of jam dicts
jams_lock = threading.Lock()

def push_alert(alert):
    try:
        with alerts_lock:
            alerts_store.insert(0, alert)  # newest first
            # trim to recent N (keep reasonable)
            if len(alerts_store) > 200:
                alerts_store[:] = alerts_store[:200]
    except Exception:
        pass

def list_alerts(limit=100):
    with alerts_lock:
        return alerts_store[:limit]

def clear_alerts():
    with alerts_lock:
        alerts_store.clear()

def push_jam(jam):
    try:
        with jams_lock:
            jams_store.insert(0, jam)
            # remove old
            cutoff = datetime.utcnow() - timedelta(seconds=JAM_RETENTION_S)
            jams_store[:] = [j for j in jams_store if datetime.fromisoformat(j["ts"]) >= cutoff]
            if len(jams_store) > 200:
                jams_store[:] = jams_store[:200]
    except Exception:
        pass

def list_jams(limit=100):
    with jams_lock:
        return jams_store[:limit]

def clear_jams():
    with jams_lock:
        jams_store.clear()

# -------------------------
# Helpers: geodesy + relative computations
# -------------------------
def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000.0
    dlat = radians(lat2-lat1)
    dlon = radians(lon2-lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1))*cos(radians(lat2)) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    return R * c

def bearing_between(lat1, lon1, lat2, lon2):
    dLon = radians(lon2 - lon1)
    lat1r = radians(lat1)
    lat2r = radians(lat2)
    x = sin(dLon) * cos(lat2r)
    y = cos(lat1r) * sin(lat2r) - sin(lat1r) * cos(lat2r) * cos(dLon)
    br = degrees(atan2(x, y))
    return (br + 360) % 360

def angle_diff(a, b):
    d = (a - b + 540) % 360 - 180
    return abs(d)

def classify_direction(bearing_self, bearing_other):
    diff = angle_diff(bearing_self, bearing_other)
    if diff < 45:
        return "same"
    if diff > 135:
        return "opposite"
    return "cross"

def closing_speed_mps(lat1, lon1, v1, bearing1, lat2, lon2, v2, bearing2):
    bearing_line = radians(bearing_between(lat1, lon1, lat2, lon2))
    theta1 = abs((bearing1 - degrees(bearing_line) + 540) % 360 - 180)
    theta2 = abs((bearing2 - (degrees(bearing_line)+180) + 540) % 360 - 180)
    proj1 = v1 * cos(radians(theta1))
    proj2 = v2 * cos(radians(theta2))
    return proj1 + proj2

def estimate_overtake_time_mps(your_speed_mps, target_speed_mps):
    dist = VEHICLE_LENGTH_M + OVERTAKE_EXTRA_M
    rel = max(0.1, your_speed_mps - target_speed_mps)
    return dist / rel

# -------------------------
# New: Confidence + Decision logic (cloud authoritative)
# -------------------------
def _recent_snapshots_for_device(device_id, limit=5):
    return Snapshot.query.filter_by(device_id=device_id).order_by(Snapshot.ts.desc()).limit(limit).all()

def _smoothed_speed_and_bearing(device_id):
    snaps = _recent_snapshots_for_device(device_id, limit=5)
    if not snaps:
        return None, None
    spd = sum([s.speed_mps for s in snaps]) / len(snaps)
    xs = sum([cos(radians(s.bearing_deg)) for s in snaps])
    ys = sum([sin(radians(s.bearing_deg)) for s in snaps])
    avg_bearing = (degrees(atan2(ys, xs)) + 360) % 360
    return float(spd), float(avg_bearing)

def _devices_near_point(lat, lon, radius_m):
    cutoff = datetime.utcnow() - timedelta(seconds=CLEANUP_STALE_SECONDS)
    snaps = Snapshot.query.filter(Snapshot.ts >= cutoff).all()
    res = {}
    for s in snaps:
        d = haversine_m(lat, lon, s.lat, s.lon)
        if d <= radius_m:
            res[s.device_id] = s
    return res

def compute_warning(self_lat, self_lon, self_speed_mps, self_bearing,
                    other_lat, other_lon, other_speed_mps, other_bearing):
    d = haversine_m(self_lat, self_lon, other_lat, other_lon)
    cls = classify_direction(self_bearing, other_bearing)
    close = closing_speed_mps(self_lat, self_lon, self_speed_mps, self_bearing,
                              other_lat, other_lon, other_speed_mps, other_bearing)
    ttc = float('inf')
    if close > 0.1:
        ttc = d / close
    guidance = {}
    guidance['distance_m'] = round(d, 2)
    guidance['direction'] = cls
    guidance['closing_mps'] = round(close, 2)
    guidance['time_to_collision_s'] = round(ttc, 2) if ttc != float('inf') else None
    if cls == "same":
        required = estimate_overtake_time_mps(self_speed_mps, other_speed_mps)
        guidance['overtake_time_required_s'] = round(required, 2)
    elif cls == "opposite":
        guidance['overtake_time_required_s'] = None
        if ttc != float('inf'):
            guidance['unsafe_if_overtaking'] = (ttc < UNSAFE_TTC_SECONDS)
    return guidance

def classify_risk(self_snap, other_snap):
    """
    Returns decision in 'red' / 'orange' / 'green' (we use 'orange' instead of 'yellow' per user's colors).
    """
    if not self_snap or not other_snap:
        return {"decision": "orange", "confidence": 0.2, "reason": "missing_data"}

    d = haversine_m(self_snap.lat, self_snap.lon, other_snap.lat, other_snap.lon)
    direction = classify_direction(self_snap.bearing_deg, other_snap.bearing_deg)
    close = closing_speed_mps(self_snap.lat, self_snap.lon, self_snap.speed_mps, self_snap.bearing_deg,
                              other_snap.lat, other_snap.lon, other_snap.speed_mps, other_snap.bearing_deg)
    ttc = float('inf')
    if close > 0.05:
        ttc = d / close

    age_self = (datetime.utcnow() - self_snap.ts).total_seconds()
    age_other = (datetime.utcnow() - other_snap.ts).total_seconds()
    recency_score = max(0.0, 1.0 - max(age_self, age_other) / max(1.0, CLEANUP_STALE_SECONDS))
    dist_score = max(0.0, 1.0 - (d / max(1.0, NEARBY_DEFAULT_RADIUS_M)))

    mid_lat = (self_snap.lat + other_snap.lat) / 2.0
    mid_lon = (self_snap.lon + other_snap.lon) / 2.0
    supporters = _devices_near_point(mid_lat, mid_lon, CONFIRMATION_RADIUS_M)
    supporters = {k:v for k,v in supporters.items() if k not in (self_snap.device_id, other_snap.device_id)}
    support_count = len(supporters)
    support_score = min(1.0, support_count / 3.0)

    base_confidence = 0.35 * recency_score + 0.40 * dist_score + 0.25 * support_score
    base_confidence = max(0.0, min(1.0, base_confidence))

    if direction == "opposite":
        if ttc == float('inf'):
            return {"decision": "green", "confidence": base_confidence * 0.6, "reason": "opposite_no_closing"}
        if ttc < UNSAFE_TTC_SECONDS:
            conf = min(1.0, base_confidence + 0.25 * (1.0 - (ttc / UNSAFE_TTC_SECONDS)) + 0.1 * support_score)
            return {"decision": "red", "confidence": round(conf, 2), "reason": f"opposite_ttc_{round(ttc,1)}s"}
        else:
            conf = base_confidence * (0.6 + 0.4 * max(0.0, (NEARBY_DEFAULT_RADIUS_M - d) / NEARBY_DEFAULT_RADIUS_M))
            return {"decision": "green", "confidence": round(conf, 2), "reason": f"opposite_ttc_safe_{round(ttc,1)}s"}

    if direction == "same":
        required = estimate_overtake_time_mps(self_snap.speed_mps, other_snap.speed_mps)
        if self_snap.speed_mps <= other_snap.speed_mps + 0.01:
            return {"decision": "green", "confidence": base_confidence * 0.6, "reason": "same_no_overtake_possible"}
        cutoff = datetime.utcnow() - timedelta(seconds=CLEANUP_STALE_SECONDS)
        other_snaps = Snapshot.query.filter(Snapshot.device_id != self_snap.device_id, Snapshot.ts >= cutoff).all()
        imminent_opposing = False
        for s in other_snaps:
            dir_s = classify_direction(self_snap.bearing_deg, s.bearing_deg)
            if dir_s == "opposite":
                d2 = haversine_m(self_snap.lat, self_snap.lon, s.lat, s.lon)
                close2 = closing_speed_mps(self_snap.lat, self_snap.lon, self_snap.speed_mps, self_snap.bearing_deg,
                                           s.lat, s.lon, s.speed_mps, s.bearing_deg)
                if close2 > 0.05:
                    ttc2 = d2 / close2
                    if ttc2 < max(UNSAFE_TTC_SECONDS, required * SAFETY_FACTOR):
                        imminent_opposing = True
                        break
        if imminent_opposing:
            conf = min(1.0, base_confidence + 0.2 * support_score)
            return {"decision": "red", "confidence": round(conf, 2), "reason": f"same_opposing_imminent_req_{round(required,1)}s"}
        if (d / max(1.0, self_snap.speed_mps*3.6)) < (required + 2.0):
            return {"decision": "orange", "confidence": round(base_confidence * 0.6 + 0.2 * support_score, 2), "reason": "same_gap_low"}
        return {"decision": "green", "confidence": round(base_confidence * 0.8 + 0.1 * support_score, 2), "reason": "same_safe"}

    if direction == "cross":
        if ttc != float('inf') and ttc < (UNSAFE_TTC_SECONDS * 0.8):
            conf = min(1.0, base_confidence + 0.15 * support_score)
            return {"decision": "red", "confidence": round(conf, 2), "reason": f"cross_ttc_{round(ttc,1)}s"}
        else:
            return {"decision": "orange", "confidence": round(base_confidence * 0.6 + 0.1 * support_score, 2), "reason": "cross_caution"}

    return {"decision": "orange", "confidence": round(base_confidence, 2), "reason": "fallback_uncertain"}

# -------------------------
# Accident detection (server-side inference)
# -------------------------
def detect_accident_for_device(device_id):
    """
    Infer a possible accident from recent snapshots for device_id.
    Returns None or an alert dict:
      {
        "accident": True,
        "severity": "high"|"medium"|"low",
        "confidence": 0.0..1.0,
        "reason": "...",
        "ts": isoformat-of-latest-snap
      }
    Uses only existing heartbeat fields (speed_mps, bearing_deg, timestamps).
    """
    snaps = _recent_snapshots_for_device(device_id, limit=8)
    if not snaps or len(snaps) < 2:
        return None

    # order ascending (oldest -> newest) for deltas
    snaps = list(reversed(snaps))
    latest = snaps[-1]
    latest_ts = latest.ts
    # only consider recent window
    window_start = datetime.utcnow() - timedelta(seconds=ACCIDENT_TIME_WINDOW_S)
    relevant = [s for s in snaps if s.ts >= window_start]
    if not relevant or len(relevant) < 2:
        relevant = snaps  # fallback to whatever we have

    # compute decelerations and bearing jumps between successive pairs
    decel_events = []
    bearing_jumps = []
    speed_drops = []
    for i in range(1, len(relevant)):
        prev = relevant[i-1]
        cur = relevant[i]
        dt = (cur.ts - prev.ts).total_seconds()
        if dt <= 0:
            continue
        dv = cur.speed_mps - prev.speed_mps
        accel = dv / dt  # m/s^2 (negative = deceleration)
        decel_events.append(accel)
        # bearing jump
        bd = angle_diff(prev.bearing_deg or 0.0, cur.bearing_deg or 0.0)
        bearing_jumps.append(bd)
        # speed drop magnitude
        drop = max(0.0, prev.speed_mps - cur.speed_mps)
        speed_drops.append((drop, dt))

    max_decel = min(decel_events) if decel_events else 0.0  # most negative (largest decel)
    max_bearing_jump = max(bearing_jumps) if bearing_jumps else 0.0
    max_speed_drop = max([s for s,d in speed_drops]) if speed_drops else 0.0
    earliest_relevant = relevant[0]

    # base confidence from how extreme the metrics are
    conf = 0.0
    reason_parts = []
    if max_decel <= -ACCIDENT_DECEL_HIGH_MPS2:
        conf += 0.5
        reason_parts.append(f"high_decel_{abs(max_decel):.1f}m/s2")
    elif max_decel <= -ACCIDENT_DECEL_MED_MPS2:
        conf += 0.3
        reason_parts.append(f"med_decel_{abs(max_decel):.1f}m/s2")

    if max_speed_drop >= ACCIDENT_SPEED_DROP_MPS:
        conf += 0.25
        reason_parts.append(f"speed_drop_{max_speed_drop:.1f}mps")

    if max_bearing_jump >= ACCIDENT_BEARING_JUMP_DEG:
        conf += 0.2
        reason_parts.append(f"bearing_jump_{int(max_bearing_jump)}deg")

    # corroborate with nearby devices that show abrupt stops within CONFIRMATION_RADIUS_M
    mid_lat = latest.lat
    mid_lon = latest.lon
    supporters = _devices_near_point(mid_lat, mid_lon, CONFIRMATION_RADIUS_M)
    corroboration = 0
    for did, s in supporters.items():
        if did == device_id:
            continue
        # check if that support device has recent snapshots showing big drop
        other_snaps = _recent_snapshots_for_device(did, limit=4)
        if len(other_snaps) >= 2:
            o_prev = other_snaps[1]
            o_last = other_snaps[0]
            dt_o = (o_last.ts - o_prev.ts).total_seconds() if (o_last.ts and o_prev.ts) else 1.0
            if dt_o > 0:
                drop_o = max(0.0, o_prev.speed_mps - o_last.speed_mps)
                if drop_o >= (ACCIDENT_SPEED_DROP_MPS * 0.6):
                    corroboration += 1
    if corroboration > 0:
        conf += min(0.2, 0.05 * corroboration)
        reason_parts.append(f"corroboration_{corroboration}")

    conf = max(0.0, min(1.0, conf))
    if conf < 0.15:
        return None

    # severity heuristics
    if conf >= 0.7 or max_decel <= -ACCIDENT_DECEL_HIGH_MPS2 or max_speed_drop >= (ACCIDENT_SPEED_DROP_MPS * 1.5):
        severity = "high"
    elif conf >= 0.4:
        severity = "medium"
    else:
        severity = "low"

    return {
        "accident": True,
        "severity": severity,
        "confidence": round(conf, 2),
        "reason": ",".join(reason_parts) if reason_parts else "inferred",
        "ts": latest_ts.isoformat() if latest_ts else None,
        "lat": latest.lat,
        "lon": latest.lon,
        "device_id": device_id
    }

# -------------------------
# DB helpers
# -------------------------
def init_db():
    with app.app_context():
        db.create_all()
        # ensure bootstrap state exists so the first admin can register even on a reused database
        try:
            _bootstrap_state_row(create_if_missing=True)
        except Exception:
            pass
        # create initial admin from env if provided and no admins exist
        try:
            if Admin.query.count() == 0 and ADMIN_USER and ADMIN_PASS:
                h = generate_password_hash(ADMIN_PASS)
                a = Admin(username=ADMIN_USER, password_hash=h)
                db.session.add(a)
                db.session.commit()
                _close_bootstrap()
                app.logger.info("Admin user created from environment variable.")
        except Exception:
            pass

def create_device_token():
    return uuid.uuid4().hex

def cleanup_old_snapshots():
    cutoff = datetime.utcnow() - timedelta(seconds=CLEANUP_STALE_SECONDS)
    Snapshot.query.filter(Snapshot.ts < cutoff).delete()
    db.session.commit()


@app.route('/')
def index():
    # Presentation-friendly landing path.
    role = _current_role()
    if role == 'admin':
        return redirect(url_for('dashboard'))
    if role == 'police':
        return redirect(url_for('police_dashboard'))
    if role == 'gk':
        return redirect(url_for('gk_dashboard'))
    return redirect(url_for('admin_login'))


# -------------------------
# Web auth helpers
# -------------------------
# Hardcoded bootstrap admin for first access.
BOOTSTRAP_ADMIN_USERNAME = 'bootstrap_admin'
BOOTSTRAP_ADMIN_PASSWORD = 'Beacon@2026!'

def _current_role():
    return session.get('auth_role') or ('admin' if session.get('admin_logged') else None)


def _set_auth_session(username, role):
    session.clear()
    session['user_id'] = username
    session['username'] = username
    session['auth_role'] = role
    session['admin_logged'] = (role == 'admin')
    session['admin_user'] = username if role == 'admin' else None
    session['police_logged'] = (role == 'police')
    session['gk_logged'] = (role == 'gk')



def _safe_render(template, **ctx):
    try:
        return render_template_string(template, **ctx)
    except Exception as exc:
        app.logger.exception("Template render failed")
        return render_template_string("""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width, initial-scale=1'><title>Beacon</title><style>body{font-family:Arial,sans-serif;background:#f6f8fb;margin:0;padding:24px}.card{max-width:860px;margin:0 auto;background:#fff;border:1px solid #e2e8f0;border-radius:14px;padding:20px;box-shadow:0 10px 24px rgba(0,0,0,.05)}a{display:inline-block;margin-top:12px;padding:10px 14px;border-radius:10px;background:#0b84ff;color:#fff;text-decoration:none;font-weight:700}</style></head><body><div class='card'><h1>Beacon</h1><p>The page loaded its core logic, but a template piece failed to render safely.</p><p><code>{{ exc }}</code></p><a href='{{ url_for("admin_login") }}'>Go to login</a></div></body></html>""", exc=str(exc))

def _pick_login_user(username):
    admin = Admin.query.filter_by(username=username).first()
    if admin:
        return admin, 'admin'
    police = PoliceUser.query.filter_by(username=username).first()
    if police:
        return police, 'police'
    gk = GKUser.query.filter_by(username=username).first()
    if gk:
        return gk, 'gk'
    return None, None

def _create_account(username, password, role='admin'):
    username = (username or '').strip()
    password = password or ''
    if not username or not password:
        return None, "Username and password are required"
    if (
        Admin.query.filter_by(username=username).first()
        or PoliceUser.query.filter_by(username=username).first()
        or GKUser.query.filter_by(username=username).first()
    ):
        return None, "Username already exists"
    pw_hash = generate_password_hash(password)
    if role == 'police':
        obj = PoliceUser(username=username, password_hash=pw_hash)
    elif role == 'gk':
        obj = GKUser(username=username, password_hash=pw_hash)
    else:
        obj = Admin(username=username, password_hash=pw_hash)
    db.session.add(obj)
    db.session.commit()
    return obj, None

def _vehicle_type_label(device):
    raw = " ".join([str(device.car_name or ""), str(device.car_model or ""), str(device.extra or "")]).lower()
    if any(k in raw for k in ('truck', 'lorry', 'hgv')):
        return 'truck'
    if any(k in raw for k in ('bus', 'coach', 'matatu')):
        return 'bus'
    if any(k in raw for k in ('motor', 'bike', 'boda')):
        return 'motorcycle'
    if any(k in raw for k in ('van', 'minibus')):
        return 'van'
    if any(k in raw for k in ('pickup', 'p-up')):
        return 'pickup'
    if any(k in raw for k in ('taxi', 'cab')):
        return 'taxi'
    return 'car'

def _zone_matches(zone, lat, lon):
    if not zone:
        return True
    if zone.scope == 'national':
        return True
    if zone.center_lat is None or zone.center_lon is None or zone.radius_m is None:
        return False
    return haversine_m(lat, lon, zone.center_lat, zone.center_lon) <= float(zone.radius_m)

def _traffic_snapshot(zone=None):
    # Use the latest known snapshot for each device, so registered vehicles still appear
    # even when they have not sent a heartbeat in the last few seconds.
    latest = {}
    for snap in Snapshot.query.order_by(Snapshot.ts.desc()).all():
        if snap.device_id not in latest:
            latest[snap.device_id] = snap

    vehicles = []
    counts = defaultdict(int)
    for d in Device.query.all():
        snap = latest.get(d.id)
        last_snapshot = {
            "ts": snap.ts.isoformat() if snap and snap.ts else None,
            "lat": snap.lat if snap else None,
            "lon": snap.lon if snap else None,
            "speed_mps": round(snap.speed_mps or 0.0, 3) if snap else None,
            "bearing_deg": round(snap.bearing_deg or 0.0, 1) if snap else None,
        }

        if snap and zone and not _zone_matches(zone, snap.lat, snap.lon):
            continue

        vehicle_type = _vehicle_type_label(d)
        if snap:
            counts[vehicle_type] += 1

        vehicles.append({
            "device_id": d.id,
            "owner": d.owner,
            "car_name": d.car_name,
            "car_model": d.car_model,
            "plate": d.plate,
            "vehicle_type": vehicle_type,
            "last_snapshot": last_snapshot,
        })

    vehicles.sort(key=lambda x: x.get("last_snapshot", {}).get("ts") or "", reverse=True)
    return vehicles, counts

def _speeders_snapshot(min_speed_kmh=80.0, zone=None):
    vehicles, _ = _traffic_snapshot(zone=zone)
    speeders = []
    for v in vehicles:
        last = v.get("last_snapshot") or {}
        spd = last.get("speed_mps")
        if spd is None:
            continue
        speed_kmh = float(spd) * 3.6
        if speed_kmh >= float(min_speed_kmh):
            item = dict(v)
            item["speed_kmh"] = round(speed_kmh, 1)
            speeders.append(item)
    speeders.sort(key=lambda x: x.get("speed_kmh", 0.0), reverse=True)
    return speeders


def _safe_login_redirect(role, next_path=None):
    if next_path and isinstance(next_path, str) and next_path.startswith('/'):
        if role == 'admin' and (next_path.startswith('/dashboard') or next_path.startswith('/friendly') or next_path.startswith('/admin/') or next_path.startswith('/report/') or next_path.startswith('/traffic')):
            return redirect(next_path)
        if role in {'admin', 'police'} and (next_path.startswith('/police') or next_path.startswith('/all-vehicles')):
            return redirect(next_path)
        if role == 'gk' and (next_path.startswith('/GK') or next_path.startswith('/gk')):
            return redirect(next_path)
    if role == 'admin':
        return redirect(url_for('dashboard'))
    if role == 'police':
        return redirect(url_for('police_dashboard'))
    if role == 'gk':
        return redirect(url_for('gk_dashboard'))
    return redirect(url_for('admin_login'))

@app.before_request
def _guard_web_pages():
    path = request.path or '/'

    public_paths = {
        '/', '/health', '/login', '/register',
        '/admin/login', '/admin/register',
        '/gk', '/GK', '/gk/login', '/GK/login',
        '/logout', '/admin/logout', '/gk/logout', '/GK/logout',
    }
    public_prefixes = ('/static/', '/socket.io/', '/onboard', '/reconnect', '/heartbeat', '/nearby', '/ingest/')
    if path in public_paths or path.startswith(public_prefixes):
        return None

    role = _current_role()
    if path.startswith(('/dashboard', '/friendly', '/admin', '/report', '/watch', '/traffic')):
        if path.startswith(('/admin/messages', '/admin/message/')):
            if role not in {'admin', 'gk'}:
                return redirect(url_for('gk_login', next=path))
        elif role != 'admin':
            return redirect(url_for('admin_login', next=path))
    elif path.startswith(('/police', '/all-vehicles')):
        if role not in {'admin', 'police'}:
            return redirect(url_for('admin_login', next=path))
    elif path.startswith(('/GK', '/gk')):
        if role not in {'admin', 'gk'}:
            return redirect(url_for('gk_login', next=path))
    return None

# -------------------------
# Authentication helpers

# -------------------------
from functools import wraps

def require_auth_token():
    """
    Traditional strict check: aborts if token missing/invalid.
    Keep for endpoints where recreation isn't wanted.
    """
    token = None
    auth = request.headers.get("Authorization")
    if auth and auth.lower().startswith("token "):
        token = auth.split(" ", 1)[1].strip()
    if not token:
        body = request.get_json(silent=True) or {}
        token = body.get("token") or request.args.get("token")
    if not token:
        abort(401, "Missing token")
    device = Device.query.filter_by(token=token, revoked=False).first()
    if not device:
        abort(401, "Invalid or revoked token")
    return device

def find_device_by_token(token):
    if not token:
        return None
    return Device.query.filter_by(token=token, revoked=False).first()

def restore_device_if_missing(device_id, token, payload=None):
    """
    If a Device row with (device_id, token) doesn't exist, recreate it using optional payload fields.
    Returns the Device instance (existing or newly created).
    """
    if not device_id or not token:
        return None
    existing = Device.query.filter_by(id=device_id).first()
    if existing:
        # if token mismatches, avoid clobbering — only create if token matches or no token present
        if existing.token != token:
            # token mismatch -> do not restore automatically
            return None
        return existing
    # create new Device row with any provided meta fields
    owner = None
    car_name = None
    car_model = None
    plate = None
    extra = None
    if payload:
        owner = payload.get("owner")
        car_name = payload.get("car_name") or payload.get("vehicle_make") or payload.get("vehicle_type")
        car_model = payload.get("car_model") or payload.get("vehicle_model_name") or payload.get("vehicle_category")
        plate = payload.get("plate")
        extra = payload.get("extra")
        if extra is None:
            extra = {}
        if isinstance(extra, dict):
            extra.setdefault("vehicle_make", payload.get("vehicle_make") or car_name)
            extra.setdefault("vehicle_model_name", payload.get("vehicle_model_name") or car_model)
            extra.setdefault("vehicle_type", payload.get("vehicle_type") or car_name)
            extra.setdefault("vehicle_category", payload.get("vehicle_category") or car_model)
            extra.setdefault("steering", payload.get("steering"))
    try:
        d = Device(id=device_id, token=token, owner=owner, car_name=car_name, car_model=car_model, plate=plate, extra=(json.dumps(extra) if extra is not None else None))
        db.session.add(d)
        db.session.commit()
        app.logger.info("Restored Device row for %s using token (auto-restore).", device_id)
        return d
    except Exception:
        db.session.rollback()
        return None

def find_or_restore_from_request(body=None):
    """
    Helper to extract token & device_id from headers or body and return device, possibly restoring the Device row.
    """
    body = body or (request.get_json(silent=True) or {})
    auth = request.headers.get("Authorization", "")
    token = None
    if auth and auth.lower().startswith("token "):
        token = auth.split(" ", 1)[1].strip()
    if not token:
        token = body.get("token") or request.args.get("token")
    device_id = body.get("device_id") or request.args.get("device_id")
    if not token:
        return None
    device = find_device_by_token(token)
    if device:
        return device
    # attempt restore if device_id present
    if device_id:
        return restore_device_if_missing(device_id, token, payload=body)
    return None

def require_admin_api():
    """
    API-level admin check: allow if logged in via session OR provide ADMIN_API_TOKEN in Bearer header.
    Use this for JSON admin endpoints. UI routes rely on session.
    """
    if session.get('admin_logged') or session.get('auth_role') == 'admin':
        return True
    auth = request.headers.get("Authorization", "")
    if auth and auth.lower().startswith("bearer "):
        token = auth.split(" ", 1)[1].strip()
        if ADMIN_API_TOKEN and token == ADMIN_API_TOKEN:
            return True
    abort(401, "Admin access required")

# -------------------------
# In-memory socket tracking
# -------------------------
connected_sockets = {}  # { device_id: set(sid) }

def send_ws_to_device(device_id, event, payload):
    sids = connected_sockets.get(device_id)
    if not sids:
        app.logger.debug("No connected sids for device %s", device_id)
        return False
    for sid in list(sids):
        try:
            socketio.emit(event, payload, room=sid)
        except Exception:
            sids.discard(sid)
    if not sids:
        connected_sockets.pop(device_id, None)
    return True

# -------------------------
# Overspeed detection helpers (new)
# -------------------------
def check_overspeed_for_snapshot(snap):
    """
    Given a Snapshot instance (persisted), check against all roads and create OverspeedEvent entries
    for any road whose circular area contains the snapshot and whose speed limit is exceeded.
    Deduplication: avoid creating multiple events for the same snapshot_id + road_id.
    """
    try:
        # convert snapshot speed to km/h for comparison
        try:
            speed_kmh = float(snap.speed_mps) * 3.6
        except Exception:
            speed_kmh = 0.0

        # only proceed if roads exist
        roads = Road.query.all()
        if not roads:
            return

        for road in roads:
            # if road has center/radius defined, use that; otherwise skip (user must define area)
            if road.center_lat is None or road.center_lon is None or road.radius_m is None:
                continue
            d = haversine_m(snap.lat, snap.lon, road.center_lat, road.center_lon)
            if d <= float(road.radius_m):
                # snapshot is inside road area
                if speed_kmh > float(road.speed_limit_kmh):
                    # dedupe: check if an event for this snapshot & road exists
                    exists = OverspeedEvent.query.filter_by(snapshot_id=snap.id, road_id=road.id).first()
                    if exists:
                        # already recorded
                        continue
                    ev = OverspeedEvent(
                        device_id=snap.device_id,
                        road_id=road.id,
                        snapshot_id=snap.id,
                        ts=snap.ts,
                        speed_kmh=round(speed_kmh, 2),
                        lat=snap.lat,
                        lon=snap.lon,
                        raw=(snap.raw)
                    )
                    db.session.add(ev)
                    try:
                        db.session.commit()
                    except Exception:
                        db.session.rollback()
                    # push to in-memory alerts list (optional)
                    try:
                        push_alert({
                            "type": "overspeed",
                            "road_id": road.id,
                            "road_name": road.name,
                            "device_id": snap.device_id,
                            "speed_kmh": round(speed_kmh, 2),
                            "lat": snap.lat,
                            "lon": snap.lon,
                            "ts": snap.ts.isoformat()
                        })
                    except Exception:
                        pass
                    # send websocket to device for feedback (best-effort)
                    try:
                        send_ws_to_device(snap.device_id, "overspeed_alert", {
                            "road_id": road.id,
                            "road_name": road.name,
                            "speed_kmh": round(speed_kmh, 2),
                            "ts": snap.ts.isoformat(),
                            "lat": snap.lat,
                            "lon": snap.lon
                        })
                    except Exception:
                        pass
    except Exception:
        app.logger.exception("check_overspeed_for_snapshot error")

# -------------------------
# Routes: API (onboard, heartbeat, nearby, admin)
# -------------------------
@app.route("/health")
def health():
    return jsonify({"ok": True, "time": datetime.utcnow().isoformat()})

@app.route("/onboard", methods=["POST"])
def onboard():
    payload = request.get_json(force=True, silent=True) or {}
    owner = payload.get("owner")
    # prefer explicit vehicle fields if provided (backwards compatible)
    car_name = payload.get("car_name") or payload.get("vehicle_make") or payload.get("vehicle_type")
    car_model = payload.get("car_model") or payload.get("vehicle_model_name") or payload.get("vehicle_category")
    plate = payload.get("plate")
    extra = payload.get("extra")
    if extra is None:
        extra = {}
    if isinstance(extra, dict):
        extra.setdefault("vehicle_make", payload.get("vehicle_make") or car_name)
        extra.setdefault("vehicle_model_name", payload.get("vehicle_model_name") or car_model)
        extra.setdefault("vehicle_type", payload.get("vehicle_type") or car_name)
        extra.setdefault("vehicle_category", payload.get("vehicle_category") or car_model)
        extra.setdefault("steering", payload.get("steering"))
    device_id = uuid.uuid4().hex
    token = create_device_token()
    d = Device(
        id=device_id,
        token=token,
        owner=owner,
        car_name=car_name,
        car_model=car_model,
        plate=plate,
        extra=(json.dumps(extra) if extra is not None else None)
    )
    db.session.add(d)
    db.session.commit()
    return jsonify({"device_id": device_id, "token": token})

@app.route("/reconnect", methods=["POST"])
def reconnect():
    """
    Optional explicit reconnect endpoint devices can call to re-create their Device row if the backend forgot it.
    Accepts device_id & token plus optional meta fields.
    """
    payload = request.get_json(force=True, silent=True) or {}
    device_id = payload.get("device_id")
    token = payload.get("token")
    if not device_id or not token:
        return jsonify({"error": "device_id and token required"}), 400
    device = find_device_by_token(token)
    if device:
        return jsonify({"ok": True, "note": "device already exists"})
    restored = restore_device_if_missing(device_id, token, payload=payload)
    if restored:
        return jsonify({"ok": True, "note": "device restored"})
    return jsonify({"error": "could not restore device"}), 500

@app.route("/heartbeat", methods=["POST"])
def heartbeat():
    # Attempt to find or restore device from Authorization header or body (so devices that kept tokens auto-recreate)
    body = request.get_json(force=True, silent=True) or {}
    device = find_or_restore_from_request(body)
    if not device:
        # fall back to strict check which will abort
        try:
            device = require_auth_token()
        except Exception:
            return jsonify({"error": "Missing or invalid token; cannot authenticate device"}), 401

    payload = body
    device_id = payload.get("device_id") or device.id

    # Rate-limiting (basic)
    now_ts = time.time()
    last = _last_heartbeat_at.get(device_id)
    if last and (now_ts - last) < HEARTBEAT_MIN_INTERVAL_S:
        return jsonify({"ok": True, "saved_at": None, "note": "rate_limited"}), 202
    _last_heartbeat_at[device_id] = now_ts

    try:
        lat = float(payload.get("lat"))
        lon = float(payload.get("lon"))
    except Exception:
        return jsonify({"error": "lat & lon required and must be numbers"}), 400
    if not (-90.0 <= lat <= 90.0 and -180.0 <= lon <= 180.0):
        return jsonify({"error": "lat/lon out of range"}), 400

    speed_mps = payload.get("speed_mps")
    if speed_mps is None:
        speed_kmh = payload.get("speed_kmh")
        if speed_kmh is not None:
            try:
                speed_mps = float(speed_kmh) / 3.6
            except Exception:
                speed_mps = 0.0
        else:
            speed_mps = 0.0
    try:
        bearing = float(payload.get("bearing", 0.0))
    except Exception:
        bearing = 0.0
    try:
        heading = float(payload.get("heading", bearing))
    except Exception:
        heading = bearing
    src = payload.get("source", "app")

    # If the device was restored with minimal meta earlier, optionally update meta if provided
    try:
        update_meta = False
        if payload.get("owner") or payload.get("car_name") or payload.get("car_model") or payload.get("plate") or payload.get("extra") or payload.get("vehicle_type") or payload.get("vehicle_category"):
            update_meta = True
        if update_meta:
            try:
                drow = Device.query.get(device.id)
                if drow:
                    if payload.get("owner"): drow.owner = payload.get("owner")
                    # support both the older keys and newer vehicle_type/category keys
                    if payload.get("car_name"): drow.car_name = payload.get("car_name")
                    if payload.get("car_model"): drow.car_model = payload.get("car_model")
                    if payload.get("vehicle_type") and not payload.get("car_name"): drow.car_name = payload.get("vehicle_type")
                    if payload.get("vehicle_category") and not payload.get("car_model"): drow.car_model = payload.get("vehicle_category")
                    if payload.get("plate"): drow.plate = payload.get("plate")
                    if payload.get("extra"): drow.extra = json.dumps(payload.get("extra"))
                    db.session.commit()
            except Exception:
                db.session.rollback()
    except Exception:
        pass

    snap = Snapshot(
        device_id=device_id,
        ts=datetime.utcnow(),
        lat=lat,
        lon=lon,
        speed_mps=float(speed_mps),
        bearing_deg=(float(bearing) % 360.0),
        heading_deg=(float(heading) % 360.0),
        source=str(src),
        raw=json.dumps(payload)
    )
    db.session.add(snap)
    db.session.commit()

    # update in-memory cache for active devices and prune old entries
    try:
        update_active_device_from_snapshot(snap)
        prune_active_devices()
    except Exception:
        app.logger.exception("active_devices update error")

    cleanup_old_snapshots()

    # compute authoritative nearby + decisions and push via socket
    try:
        nearby_payload = compute_nearby_for_device(device_id, radius_m=NEARBY_DEFAULT_RADIUS_M)

        # Accident detection for the reporting device (server-side inferred)
        try:
            acc = detect_accident_for_device(device_id)
            if acc:
                # attach alerts into nearby payload
                nearby_payload.setdefault("alerts", []).append(acc)
                # push device-specific accident alert
                send_ws_to_device(device_id, 'accident_alert', acc)
                # record in global alerts store
                try:
                    push_alert(acc)
                except Exception:
                    pass
                # push a minimal "accident_nearby" to nearby connected devices so they can react
                for other in nearby_payload.get("nearby", []):
                    try:
                        send_ws_to_device(other["device_id"], 'accident_nearby', {
                            "accident": True,
                            "device_id": device_id,
                            "lat": acc.get("lat"),
                            "lon": acc.get("lon"),
                            "severity": acc.get("severity"),
                            "confidence": acc.get("confidence"),
                            "reason": acc.get("reason"),
                            "ts": acc.get("ts")
                        })
                    except Exception:
                        # continue best-effort
                        pass
        except Exception:
            app.logger.exception("accident detection error")

        send_ws_to_device(device_id, 'nearby_update', nearby_payload)
    except Exception as e:
        app.logger.exception("compute_nearby error: %s", e)

    # NEW: after persisting snapshot, check for overspeed (per-road)
    try:
        check_overspeed_for_snapshot(snap)
    except Exception:
        app.logger.exception("overspeed check error")

    return jsonify({"ok": True, "saved_at": snap.ts.isoformat()})

# -------------------------
# Nearby & compute_for_device functions (unchanged)
# -------------------------
@app.route("/nearby", methods=["GET"])
def nearby():
    token_device = None
    try:
        token_device = require_auth_token()
    except Exception:
        token_device = None
    device_id = request.args.get("device_id") or (token_device.id if token_device else None)
    if not device_id:
        return jsonify({"error": "device_id required (or provide Authorization token)"}), 400
    radius_m = float(request.args.get("radius_m", NEARBY_DEFAULT_RADIUS_M))
    try:
        payload = compute_nearby_for_device(device_id, radius_m=radius_m)
        # include server-side accident inference when requested via HTTP as well
        try:
            acc = detect_accident_for_device(device_id)
            if acc:
                payload.setdefault("alerts", []).append(acc)
        except Exception:
            pass
        return jsonify(payload)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def compute_nearby_for_device(device_id, radius_m=NEARBY_DEFAULT_RADIUS_M):
    # obtain the latest snapshot for the device (or fallback to active_devices cache)
    self_snap = Snapshot.query.filter_by(device_id=device_id).order_by(Snapshot.ts.desc()).first()
    if not self_snap:
        with active_devices_lock:
            entry = active_devices.get(device_id)
        if entry:
            class _Tmp:
                pass
            t = _Tmp()
            t.device_id = device_id
            t.lat = entry.get("lat")
            t.lon = entry.get("lon")
            t.speed_mps = entry.get("speed_mps")
            t.bearing_deg = entry.get("bearing_deg")
            t.ts = entry.get("ts")
            t.source = entry.get("source", "app")
            self_snap = t
        else:
            raise RuntimeError("no snapshot for device")

    # gather other recent snapshots (DB + in-memory cache)
    cutoff = datetime.utcnow() - timedelta(seconds=CLEANUP_STALE_SECONDS)
    other_snaps = Snapshot.query.filter(Snapshot.device_id != device_id, Snapshot.ts >= cutoff).all()

    with active_devices_lock:
        for did, entry in active_devices.items():
            if did == device_id:
                continue
            seen = any(s.device_id == did for s in other_snaps)
            if seen:
                continue
            if entry.get("ts") >= cutoff:
                class _Tmp2:
                    pass
                t = _Tmp2()
                t.device_id = did
                t.lat = entry.get("lat")
                t.lon = entry.get("lon")
                t.speed_mps = entry.get("speed_mps")
                t.bearing_deg = entry.get("bearing_deg")
                t.ts = entry.get("ts")
                t.source = entry.get("source", "app")
                other_snaps.append(t)

    results = []
    for s in other_snaps:
        try:
            d = haversine_m(self_snap.lat, self_snap.lon, s.lat, s.lon)
        except Exception:
            continue
        if d > radius_m:
            continue
        direction = classify_direction(self_snap.bearing_deg, s.bearing_deg)
        closing = closing_speed_mps(self_snap.lat, self_snap.lon, self_snap.speed_mps, self_snap.bearing_deg,
                                    s.lat, s.lon, s.speed_mps, s.bearing_deg)
        guidance = compute_warning(self_snap.lat, self_snap.lon, self_snap.speed_mps, self_snap.bearing_deg,
                                   s.lat, s.lon, s.speed_mps, s.bearing_deg)
        risk = classify_risk(self_snap, s)

        # try to include device metadata from Device row (best-effort)
        meta_owner = None
        meta_car_name = None
        meta_car_model = None
        meta_plate = None
        try:
            drow = Device.query.get(s.device_id)
            if drow:
                meta_owner = drow.owner
                meta_car_name = drow.car_name
                meta_car_model = drow.car_model
                meta_plate = drow.plate
        except Exception:
            pass

        results.append({
            "device_id": s.device_id,
            "ts": s.ts.isoformat() if hasattr(s, "ts") else None,
            "lat": s.lat,
            "lon": s.lon,
            "distance_m": round(d, 2),
            "direction": direction,
            "speed_mps": round(s.speed_mps, 2),
            "bearing_deg": round(s.bearing_deg, 1),
            "closing_mps": round(closing, 2),
            "guidance": guidance,
            "decision": risk.get("decision"),
            "confidence": float(risk.get("confidence", 0.0)),
            "reason": risk.get("reason"),
            # meta fields (for client convenience)
            "owner": meta_owner,
            "car_name": meta_car_name,
            "car_model": meta_car_model,
            "plate": meta_plate
        })
    # sort by distance so client shows closest first
    results.sort(key=lambda x: x["distance_m"])

    # self meta
    self_meta = {}
    try:
        dro = Device.query.get(device_id)
        if dro:
            self_meta = {
                "owner": dro.owner,
                "car_name": dro.car_name,
                "car_model": dro.car_model,
                "plate": dro.plate
            }
    except Exception:
        pass

    payload = {
        "self": {
            "device_id": self_snap.device_id,
            "lat": self_snap.lat,
            "lon": self_snap.lon,
            "speed_mps": round(self_snap.speed_mps, 2),
            "bearing_deg": round(self_snap.bearing_deg, 1),
            "ts": self_snap.ts.isoformat() if hasattr(self_snap, "ts") else None,
            "meta": self_meta
        },
        "nearby": results
    }

    # include server-inferred accidents for reporting + nearby devices (best-effort)
    alerts = []
    try:
        acc_self = detect_accident_for_device(device_id)
        if acc_self:
            alerts.append(acc_self)
        for r in results[:8]:
            try:
                a = detect_accident_for_device(r["device_id"])
                if a:
                    alerts.append(a)
            except Exception:
                pass
    except Exception:
        pass

    if alerts:
        payload["alerts"] = alerts

    return payload

# -------------------------
# Admin/UI templates and routes
# -------------------------
ADMIN_LOGIN_HTML = """
<!doctype html>
<html>
<head><meta charset="utf-8"><title>Beacon Login</title><meta name="viewport" content="width=device-width, initial-scale=1" />
<style>
body{font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,Arial;background:linear-gradient(180deg,#07111f 0%, #0b1220 100%);margin:0;padding:24px;color:#e5eefb;}
.card{max-width:560px;margin:0 auto;background:rgba(16,26,45,.96);border:1px solid rgba(148,163,184,.15);border-radius:24px;padding:24px;box-shadow:0 24px 64px rgba(0,0,0,.28);backdrop-filter: blur(12px);} 
input{width:100%;box-sizing:border-box;padding:12px;border:1px solid rgba(148,163,184,.18);border-radius:14px;margin-top:8px;background:#0b1324;color:#e5eefb;}
button,a{display:inline-block;margin-top:16px;padding:12px 16px;border-radius:999px;border:0;background:linear-gradient(135deg,#8b5cf6,#38bdf8);color:#fff;text-decoration:none;font-weight:800;cursor:pointer;}
.muted{color:#94a3b8;font-size:14px;line-height:1.5;}
.flash{background:rgba(239,68,68,.12);border:1px solid rgba(239,68,68,.25);color:#fecaca;padding:12px;border-radius:14px;margin-top:12px;}
.top{display:flex;gap:10px;flex-wrap:wrap;margin-top:14px;}
</style>
</head>
<body>
  <div class="card">
    <h2>Beacon Login</h2>
    {% with messages = get_flashed_messages() %}
      {% if messages %}
        <div class="flash">{{ messages[0] }}</div>
      {% endif %}
    {% endwith %}
    <form method="post" action="{{ url_for('admin_login') }}">
      <label>Username</label>
      <input name="username" required>
      <label>Password</label>
      <input name="password" type="password" required>
      <button type="submit">Login</button>
    </form>
    <div class="top">
      {% if allow_register %}
        <a href="{{ url_for('admin_register') }}">Register first admin</a>
      {% endif %}
      <span class="muted">Admins open the dashboard. Police users open the police dashboard. GK users open /GK.</span>
    </div>
    {% if allow_register %}
      <p class="muted">This is the very first setup. Once the first admin exists, registration disappears.</p>
    {% else %}
      <p class="muted">Registration is closed. An existing admin can add more admins or police users.</p>
    {% endif %}
  </div>
</body>
</html>
"""

# -------------------------
# Friendly Dashboard HTML (modified to include floating widget for Road & Reports)
# -------------------------
# We keep the bulk of your original DASHBOARD_HTML content intact (map, devices list, modal).
# Added a floating widget (minimizable) in the bottom-right for Road search / create / select,
# plus buttons to download reports (all / per-road) as Excel or PDF.
DASHBOARD_HTML = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Beacon — Dashboard</title>
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
  <style>
    :root{ --bg:#f7f9fb; --card:#ffffff; --muted:#6b7280; --accent:#0b84ff; }
    body{ font-family: Inter, system-ui, -apple-system, "Segoe UI", Roboto, Arial; margin:0; background:var(--bg); color:#111827;}
    header{ background: linear-gradient(90deg,#0b84ff 0%, #00c6ff 100%); color:white; padding:18px 20px; display:flex; align-items:center; gap:12px;}
    header h1{ font-size:18px; margin:0;}
    .wrap{ display:flex; gap:12px; padding:12px; height: calc(100vh - 72px); box-sizing:border-box;}
    .left{ width:360px; display:flex; flex-direction:column; gap:12px; }
    .card{ background:var(--card); border-radius:10px; padding:12px; box-shadow:0 6px 18px rgba(15,23,42,0.06); overflow:auto; }
    #devicesList{ list-style:none; margin:0; padding:0; }
    #devicesList li{ padding:10px; border-radius:8px; margin-bottom:8px; cursor:pointer; border:1px solid #eef2f7; display:flex; justify-content:space-between; align-items:center; gap:8px;}
    #devicesList li.selected{ background:#eef8ff; border-color:#cfe9ff; }
    .dev-meta{ font-size:13px; color:var(--muted); }
    .big{ font-weight:600; font-size:14px; }
    .muted{ color:var(--muted); font-size:13px; }
    #mapWrap{ flex:1; display:flex; flex-direction:column; gap:12px; }
    #map{ flex:1; border-radius:10px; overflow:hidden; }
    #detailCard{ height:220px; min-height:160px; transition: all 0.18s ease; }
    .row{ display:flex; justify-content:space-between; gap:8px; margin-top:6px; }
    .btn{ background:var(--accent); color:white; padding:8px 10px; border-radius:8px; border:none; cursor:pointer;}
    .btn.ghost{ background:transparent; color:var(--accent); border:1px solid #e6f2ff;}
    .small{ font-size:12px; padding:6px 8px; border-radius:6px; }
    a.osm{ color:var(--accent); text-decoration:none; font-weight:600; }
    footer { font-size:12px; color:var(--muted); padding:8px 16px; text-align:right; }
    #statusBar { font-size:13px; color:#08306B; margin-top:8px; }

    /* modal (expanded details) */
    .modal { position: fixed; inset: 0; display:none; align-items:center; justify-content:center; z-index:1200; }
    .modal.show { display:flex; }
    .modal-backdrop { position:absolute; inset:0; background: rgba(2,6,23,0.55); }
    .modal-window { position:relative; width:90%; max-width:1000px; height:85%; background:white; border-radius:12px; padding:16px; overflow:auto; box-shadow:0 18px 46px rgba(2,6,23,0.36); z-index:1210; }
    .modal .close { position:absolute; right:12px; top:12px; background:#eee; border:none; padding:6px 8px; border-radius:8px; cursor:pointer; }

    pre#detailRaw { background:#fbfdff; padding:12px; border-radius:6px; max-height:120px; overflow:auto; white-space:pre-wrap; word-wrap:break-word; }
    .meta-row { margin-top:8px; font-size:13px; color:#374151; }

    /* Floating roads widget */
    #roadsWidget {
      position: fixed;
      right: 16px;
      bottom: 16px;
      width: 360px;
      max-width: calc(100% - 32px);
      z-index: 1300;
      transition: transform 0.18s ease, opacity 0.18s ease;
    }
    #roadsWidget .widget-card {
      background: var(--card);
      border-radius: 12px;
      padding: 12px;
      box-shadow: 0 12px 36px rgba(2,6,23,0.24);
      overflow: hidden;
    }
    #roadsWidget .widget-header { display:flex; justify-content:space-between; align-items:center; gap:8px; }
    #roadsWidget .widget-body { margin-top:10px; max-height:340px; overflow:auto; }
    #roadsWidget.minimized .widget-body { display:none; }
    #roadsWidget .road-item { padding:8px; border-radius:8px; border:1px solid #eef2f7; margin-bottom:8px; display:flex; justify-content:space-between; align-items:center; }
    #roadsWidget .tiny { font-size:12px; color:var(--muted); }
    #roadsWidget input, #roadsWidget select { width:100%; padding:8px; margin-top:6px; border-radius:8px; border:1px solid #e6eef8; box-sizing:border-box; }
    #roadsWidget .flex-row { display:flex; gap:8px; }

    /* small responsiveness */
    @media (max-width:700px){ .left{ display:none; } #roadsWidget{ width:95%; right:8px; bottom:8px;} }
  </style>
</head>
<body>
  <header>
    <h1>Beacon — Live Dashboard</h1>
    <div style="margin-left:auto; display:flex; gap:8px; flex-wrap:wrap; align-items:center;">
      <a href="{{ url_for('all_vehicles') }}" style="color:white;text-decoration:none;font-weight:700;background:rgba(255,255,255,.15);padding:8px 12px;border-radius:10px;">All Vehicles</a>
      <a href="{{ url_for('admin_admins') }}" style="color:white;text-decoration:none;font-weight:700;background:rgba(255,255,255,.15);padding:8px 12px;border-radius:10px;">Users</a>
      <a href="{{ url_for('admin_traffic') }}" style="color:white;text-decoration:none;font-weight:700;background:rgba(255,255,255,.15);padding:8px 12px;border-radius:10px;">Traffic</a>
      <a href="{{ url_for('admin_speeders') }}" style="color:white;text-decoration:none;font-weight:700;background:rgba(255,255,255,.15);padding:8px 12px;border-radius:10px;">Speeders</a>
      <a href="{{ url_for('admin_messages') }}" style="color:white;text-decoration:none;font-weight:700;background:rgba(255,255,255,.15);padding:8px 12px;border-radius:10px;">Messages</a>
      <span style="font-size:13px; opacity:0.95;">Auto-refresh every 5s — open this page on desktop or phone</span>
    </div>
  </header>

  <div class="wrap">
    <div class="left">
      <div class="card">
        <div style="display:flex; justify-content:space-between; align-items:center;">
          <div>
            <div class="muted">Devices</div>
            <div class="big" id="devicesCount">0 devices</div>
          </div>
          <div>
            <button id="btnRefresh" class="btn small">Refresh</button>
          </div>
        </div>

        <div id="statusBar" class="muted">Status: idle</div>

        <hr style="margin:10px 0; border:none; border-top:1px solid #f1f5f9;" />
        <ul id="devicesList"></ul>
      </div>

      <div id="detailCard" class="card">
        <div id="placeholderDetail"><em>Select a device to see details & map</em></div>
        <div id="deviceDetail" style="display:none;">
          <div style="display:flex; justify-content:space-between; align-items:center;">
            <div>
              <div id="detailName" class="big"></div>
              <div id="detailOwner" class="muted"></div>
            </div>
            <div>
              <button id="btnExpand" class="btn small" title="Expand details">Expand</button>
              <button id="btnRevoke" class="btn ghost small">Revoke</button>
            </div>
          </div>

          <div class="row">
            <div><div class="muted">Last seen</div><div id="detailTs" class="big"></div></div>
            <div><div class="muted">Speed</div><div id="detailSpeed" class="big">—</div></div>
          </div>

          <div class="row">
            <div><div class="muted">Location</div><div id="detailLoc" class="muted"></div></div>
            <div><a id="osmLink" class="osm" target="_blank">Open in OSM ↗</a></div>
          </div>

          <div style="margin-top:8px;">
            <div class="muted">Raw heartbeat (last)</div>
            <pre id="detailRaw" style="background:#fbfdff; padding:8px; border-radius:6px; max-height:120px; overflow:auto;">—</pre>
          </div>
        </div>
      </div>

    </div>

    <div id="mapWrap" class="card">
      <div id="map"></div>
    </div>
  </div>

  <footer>Beacon — simple live tracking dashboard</footer>

  <!-- Expanded modal -->
  <div id="modal" class="modal" role="dialog" aria-hidden="true">
    <div class="modal-backdrop" onclick="closeModal()"></div>
    <div class="modal-window" id="modalWindow" role="document">
      <button class="close" onclick="closeModal()">Close ✕</button>
      <h2 id="modalTitle">Device details</h2>
      <div id="modalContent">
        <div class="meta-row" id="modalMeta">Loading…</div>
        <hr/>
        <div><strong>Recent snapshots</strong></div>
        <pre id="modalSnapshots" style="white-space:pre-wrap; word-break:break-word; background:#fbfdff; padding:12px; border-radius:8px; max-height:60%; overflow:auto;">Loading…</pre>
      </div>
    </div>
  </div>

  <!-- Floating Roads & Reports widget -->
  <div id="roadsWidget" class="">
    <div class="widget-card">
      <div class="widget-header">
        <div>
          <strong>Roads & Reports</strong><br/>
          <span class="tiny">Search, create, monitor, export</span>
        </div>
        <div style="display:flex; gap:8px; align-items:center;">
          <button id="toggleWidget" class="btn small">−</button>
        </div>
      </div>

      <div class="widget-body">
        <div style="margin-top:8px;">
          <input id="searchRoadInput" placeholder="Search roads by name..." />
        </div>
        <div id="roadsList" style="margin-top:8px;"></div>

        <hr/>
        <div><strong>Create new road area</strong></div>
        <div style="margin-top:8px;">
          <input id="newRoadName" placeholder="Road name (e.g. 'Mombasa Rd')" />
          <div class="flex-row" style="margin-top:8px;">
            <input id="newRoadSpeed" placeholder="Speed limit (km/h)" />
            <input id="newRoadRadius" placeholder="Radius (m)" />
          </div>
          <div style="margin-top:8px;">
            <input id="newRoadLat" placeholder="Center lat (optional)" />
            <input id="newRoadLon" placeholder="Center lon (optional)" />
          </div>
          <div style="margin-top:8px; display:flex; gap:8px;">
            <button id="btnCreateRoad" class="btn small">Create Road</button>
            <button id="btnRefreshRoads" class="btn ghost small">Refresh Roads</button>
          </div>
        </div>

        <hr/>
        <div style="display:flex; gap:8px; margin-top:6px;">
          <button id="btnExportAllXLSX" class="btn small">Export All (.xlsx)</button>
          <button id="btnExportAllPDF" class="btn small">Export All (.pdf)</button>
        </div>
        <div style="margin-top:6px;">
          <span class="tiny">Select a road then use the per-road export buttons next to it.</span>
        </div>
      </div>
    </div>
  </div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://cdn.jsdelivr.net/npm/luxon@3/build/global/luxon.min.js"></script>
<script>
  // small safety: guard usage if variable 'd' is not present
  try {
    const DateTime = luxon.DateTime;
    if (typeof d !== 'undefined' && d.last_snapshot && d.last_snapshot.ts) {
      const dt = DateTime.fromISO(d.last_snapshot.ts, { zone: 'utc' }).setZone('Africa/Nairobi');
      document.getElementById('detailTs').innerText = dt.toLocaleString(DateTime.DATETIME_MED);
    }
  } catch(e) {}
</script>
<script>
  const devicesListEl = document.getElementById('devicesList');
  const devicesCountEl = document.getElementById('devicesCount');
  const btnRefresh = document.getElementById('btnRefresh');
  const statusBar = document.getElementById('statusBar');

  const btnExpand = document.getElementById('btnExpand');
  const modal = document.getElementById('modal');
  const modalTitle = document.getElementById('modalTitle');
  const modalMeta = document.getElementById('modalMeta');
  const modalSnapshots = document.getElementById('modalSnapshots');

  let devicesCache = [];
  let selectedId = null;
  let refreshTimer = null;

  // Leaflet map init
  const map = L.map('map', { center: [0,0], zoom: 2, preferCanvas:true });
  L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', { maxZoom: 19 }).addTo(map);
  let marker = null;
  let circle = null;

  btnRefresh.addEventListener('click', fetchDevices);
  btnExpand.addEventListener('click', () => { if (selectedId) openModal(selectedId); });

  async function fetchDevices(){
    statusBar.innerText = "Status: fetching /admin/devices...";
    try {
      const res = await fetch('/admin/devices', { cache: "no-store" });
      const txt = await res.text();
      statusBar.innerText = `Status: HTTP ${res.status} — response ${txt.length} bytes`;
      let data;
      try {
        data = JSON.parse(txt);
      } catch (e) {
        console.warn("JSON.parse failed, trying fallback:", e);
        const m = txt.match(/\\{\\s*"?devices"?:[\\s\\S]*\\}/);
        if (m) {
          data = JSON.parse(m[0]);
        } else {
          throw new Error("Response not valid JSON");
        }
      }
      devicesCache = data.devices || [];
      statusBar.innerText = `Status: fetched ${devicesCache.length} device(s)`;
      renderList();
    } catch (err) {
      console.error("fetchDevices error:", err);
      statusBar.innerText = "Status: fetch error — see console for details";
      devicesCache = [];
      renderList();
    }
  }

  function renderList(){
    devicesListEl.innerHTML = '';
    const count = devicesCache.length || 0;
    devicesCountEl.innerText = (count === 1) ? "1 device" : (count + " devices");

    devicesCache.sort((a,b) => {
      const ta = a.last_snapshot && a.last_snapshot.ts ? new Date(a.last_snapshot.ts).getTime() : 0;
      const tb = b.last_snapshot && b.last_snapshot.ts ? new Date(b.last_snapshot.ts).getTime() : 0;
      return tb - ta;
    });

    for (const d of devicesCache) {
      const li = document.createElement('li');
      li.dataset.id = d.id;
      const name = d.car_name || d.car_model || d.id;
      const last = d.last_snapshot ? (d.last_snapshot.ts ? new Date(d.last_snapshot.ts).toLocaleString() : 'seen') : 'never';
      const speed = d.last_snapshot ? (((d.last_snapshot.speed_mps || 0) * 3.6).toFixed(1) + ' km/h') : '—';
      li.innerHTML = `<div style="min-width:0;">
                        <div class="big">${escapeHtml(name)}</div>
                        <div class="dev-meta">${escapeHtml(d.owner || '')} — ${escapeHtml(d.plate || '')}</div>
                      </div>
                      <div style="text-align:right;">
                        <div class="muted">${escapeHtml(last)}</div>
                        <div class="muted">${escapeHtml(speed)}</div>
                      </div>`;
      li.addEventListener('click', () => selectDevice(d.id));
      if (d.id === selectedId) li.classList.add('selected');
      devicesListEl.appendChild(li);
    }

    if (!selectedId && devicesCache.length > 0) {
      selectDevice(devicesCache[0].id);
    }
  }

  function selectDevice(id) {
    selectedId = id;
    document.querySelectorAll('#devicesList li').forEach(li => {
      li.classList.toggle('selected', li.dataset.id === id);
    });
    const d = devicesCache.find(x => x.id === id);
    if (!d) return;
    showDetail(d);
    if (d.last_snapshot && d.last_snapshot.lat && d.last_snapshot.lon) {
      placeMarker(d.last_snapshot.lat, d.last_snapshot.lon, d);
    } else {
      clearMarker();
    }
  }

  function showDetail(d) {
    document.getElementById('placeholderDetail').style.display = 'none';
    const panel = document.getElementById('deviceDetail');
    panel.style.display = 'block';
    document.getElementById('detailName').innerText = d.car_name || d.car_model || d.id;
    document.getElementById('detailOwner').innerText = d.owner || '';
    if (d.last_snapshot) {
      document.getElementById('detailTs').innerText = d.last_snapshot.ts ? new Date(d.last_snapshot.ts).toLocaleString() : 'seen';
      const spd = ((d.last_snapshot.speed_mps || 0) * 3.6).toFixed(1) + ' km/h';
      document.getElementById('detailSpeed').innerText = spd;
      document.getElementById('detailLoc').innerText = (typeof d.last_snapshot.lat === 'number' && typeof d.last_snapshot.lon === 'number')
        ? (d.last_snapshot.lat.toFixed(6) + ', ' + d.last_snapshot.lon.toFixed(6)) : '—';
      document.getElementById('osmLink').href = (d.last_snapshot && d.last_snapshot.lat && d.last_snapshot.lon)
        ? `https://www.openstreetmap.org/?mlat=${d.last_snapshot.lat}&mlon=${d.last_snapshot.lon}#map=18/${d.last_snapshot.lat}/${d.last_snapshot.lon}`
        : '#';
      document.getElementById('detailRaw').innerText = JSON.stringify(d.last_snapshot.raw || d.last_snapshot, null, 2);
      document.getElementById('btnRevoke').onclick = async () => {
        if (!confirm('Revoke device token? This prevents the app from authenticating with that token.')) return;
        try {
          const res = await fetch('/admin/device/' + d.id + '/revoke', { method: 'POST' });
          if (!res.ok) { alert('Revoke failed'); return; }
          alert('Device revoked');
          fetchDevices();
        } catch (e) { alert('Revoke error'); }
      };
    } else {
      document.getElementById('detailTs').innerText = '—';
      document.getElementById('detailSpeed').innerText = '—';
      document.getElementById('detailLoc').innerText = '—';
      document.getElementById('detailRaw').innerText = '—';
      document.getElementById('osmLink').href = '#';
    }
  }

  function placeMarker(lat, lon, d) {
    if (!marker) {
      marker = L.marker([lat, lon]).addTo(map);
    } else {
      marker.setLatLng([lat, lon]);
    }
    if (!circle) {
      circle = L.circle([lat, lon], { radius: (d.last_snapshot && d.last_snapshot.raw && d.last_snapshot.raw.accuracy) ? d.last_snapshot.raw.accuracy : 20 }).addTo(map);
    } else {
      circle.setLatLng([lat, lon]);
    }
    map.setView([lat, lon], 15, { animate: true });
    marker.bindPopup(`<strong>${escapeHtml(d.car_name || d.id)}</strong><br/>${d.last_snapshot && d.last_snapshot.ts ? new Date(d.last_snapshot.ts).toLocaleString() : ''}`).openPopup();
  }

  function clearMarker(){
    if (marker) { map.removeLayer(marker); marker = null; }
    if (circle) { map.removeLayer(circle); circle = null; }
  }

  function escapeHtml(s) {
    if (!s) return '';
    return s.replace(/[&<>"'`]/g, (c) => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;',"`":'&#96;'})[c]);
  }

  // modal functions
  async function openModal(deviceId) {
    modal.classList.add('show');
    modalTitle.innerText = 'Device details — ' + deviceId;
    modalMeta.innerText = 'Loading…';
    modalSnapshots.innerText = 'Loading…';
    try {
      const res = await fetch('/admin/device/' + deviceId + '/json', { cache: "no-store" });
      const j = await res.json();
      const dev = j.device || {};
      modalMeta.innerHTML = `
        <div><strong>Name:</strong> ${escapeHtml(dev.car_name || dev.car_model || dev.id || '')}</div>
        <div><strong>Owner:</strong> ${escapeHtml(dev.owner || '')}</div>
        <div><strong>Plate:</strong> ${escapeHtml(dev.plate || '')}</div>
        <div><strong>Created:</strong> ${escapeHtml(dev.created_at || '')}</div>
        <div><strong>Connected:</strong> ${j.connected ? 'yes' : 'no'}</div>
      `;
      modalSnapshots.innerText = JSON.stringify(j.snapshots || [], null, 2);
    } catch (err) {
      modalMeta.innerText = 'Error loading details';
      modalSnapshots.innerText = String(err);
    }
  }
  function closeModal() {
    modal.classList.remove('show');
  }

  // auto-refresh
  refreshTimer = setInterval(fetchDevices, 5000);

  // initial load
  fetchDevices();

  // -----------------------------
  // Roads widget logic (frontend)
  // -----------------------------
  const toggleWidget = document.getElementById('toggleWidget');
  const roadsWidget = document.getElementById('roadsWidget');
  const roadsList = document.getElementById('roadsList');
  const searchRoadInput = document.getElementById('searchRoadInput');
  const btnCreateRoad = document.getElementById('btnCreateRoad');
  const btnRefreshRoads = document.getElementById('btnRefreshRoads');
  const btnExportAllXLSX = document.getElementById('btnExportAllXLSX');
  const btnExportAllPDF = document.getElementById('btnExportAllPDF');

  let roadsCache = [];
  let selectedRoadId = null;

  toggleWidget.addEventListener('click', () => {
    if (roadsWidget.classList.contains('minimized')) {
      roadsWidget.classList.remove('minimized');
      toggleWidget.innerText = '−';
    } else {
      roadsWidget.classList.add('minimized');
      toggleWidget.innerText = '+';
    }
  });

  async function fetchRoads(){
    try {
      const res = await fetch('/admin/roads', { cache: "no-store" });
      const j = await res.json();
      roadsCache = j.roads || [];
      renderRoads();
    } catch (e) {
      console.error("fetchRoads error", e);
      roadsCache = [];
      renderRoads();
    }
  }

  function renderRoads(){
    const filter = (searchRoadInput.value || '').toLowerCase().trim();
    roadsList.innerHTML = '';
    const list = roadsCache.filter(r => !filter || (r.name && r.name.toLowerCase().includes(filter)));
    for (const r of list) {
      const div = document.createElement('div');
      div.className = 'road-item';
      div.innerHTML = `
        <div>
          <div style="font-weight:600">${escapeHtml(r.name)}</div>
          <div class="tiny">${escapeHtml((r.speed_limit_kmh||'') + ' km/h')} — radius ${escapeHtml((r.radius_m||'') + ' m')}</div>
        </div>
        <div style="display:flex; flex-direction:column; gap:6px; align-items:flex-end;">
          <div style="display:flex; gap:6px;">
            <button class="btn small btn-monitor" data-id="${r.id}">Monitor</button>
            <button class="btn ghost small btn-export" data-id="${r.id}">Export</button>
          </div>
          <div style="display:flex; gap:6px; margin-top:6px;">
            <button class="btn ghost small btn-delete" data-id="${r.id}">Delete</button>
          </div>
        </div>
      `;
      roadsList.appendChild(div);
    }

    // hook buttons
    document.querySelectorAll('.btn-monitor').forEach(b => {
      b.onclick = async (ev) => {
        const id = b.dataset.id;
        selectedRoadId = id;
        alert('Monitoring road selected. Overspeed will be recorded automatically for devices in that road area.');
      };
    });
    document.querySelectorAll('.btn-export').forEach(b => {
      b.onclick = async (ev) => {
        const id = b.dataset.id;
        // open per-road excel in new tab
        window.open('/report/road/' + id + '.xlsx', '_blank');
      };
    });
    document.querySelectorAll('.btn-delete').forEach(b => {
      b.onclick = async (ev) => {
        const id = b.dataset.id;
        if (!confirm('Delete road? This will remove it from monitoring history but not delete past overspeed events.')) return;
        try {
          const res = await fetch('/admin/road/' + id, { method: 'DELETE' });
          if (!res.ok) { alert('Delete failed'); return; }
          fetchRoads();
        } catch (e) { alert('Delete error'); }
      };
    });
  }

  searchRoadInput.addEventListener('input', () => renderRoads());
  btnRefreshRoads.addEventListener('click', fetchRoads);

  btnCreateRoad.addEventListener('click', async () => {
    const name = (document.getElementById('newRoadName').value || '').trim();
    const speed = parseFloat((document.getElementById('newRoadSpeed').value || '').trim());
    const radius = parseFloat((document.getElementById('newRoadRadius').value || '').trim()) || 50;
    const lat = parseFloat((document.getElementById('newRoadLat').value || '').trim());
    const lon = parseFloat((document.getElementById('newRoadLon').value || '').trim());
    if (!name || isNaN(speed)) { alert('Name and speed limit required'); return; }
    const body = { name: name, speed_limit_kmh: speed, radius_m: radius };
    if (!isNaN(lat) && !isNaN(lon)) {
      body.center_lat = lat; body.center_lon = lon;
    }
    try {
      const res = await fetch('/admin/roads', { method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify(body) });
      if (!res.ok) {
        const txt = await res.text();
        alert('Create road failed: ' + txt);
        return;
      }
      document.getElementById('newRoadName').value = '';
      document.getElementById('newRoadSpeed').value = '';
      document.getElementById('newRoadRadius').value = '';
      document.getElementById('newRoadLat').value = '';
      document.getElementById('newRoadLon').value = '';
      await fetchRoads();
    } catch (e) {
      alert('Create error: ' + e);
    }
  });

  btnExportAllXLSX.addEventListener('click', () => {
    window.open('/report/all.xlsx', '_blank');
  });
  btnExportAllPDF.addEventListener('click', () => {
    window.open('/report/all.pdf', '_blank');
  });

  // initial load for roads
  fetchRoads();

</script>
</body>
</html>
"""

# Admin web pages / endpoints

@app.route('/login', methods=['GET', 'POST'])
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'GET':
        return _safe_render(ADMIN_LOGIN_HTML, allow_register=False)

    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    if not username or not password:
        flash("Missing username or password")
        return _safe_render(ADMIN_LOGIN_HTML, allow_register=False), 400

    # Bootstrap login: lets you enter the app immediately on first use.
    # The account is created on the fly if it does not already exist.
    if username == BOOTSTRAP_ADMIN_USERNAME and password == BOOTSTRAP_ADMIN_PASSWORD:
        try:
            existing = Admin.query.filter_by(username=BOOTSTRAP_ADMIN_USERNAME).first()
            if not existing:
                existing = Admin(
                    username=BOOTSTRAP_ADMIN_USERNAME,
                    password_hash=generate_password_hash(BOOTSTRAP_ADMIN_PASSWORD),
                )
                db.session.add(existing)
                db.session.commit()
        except Exception:
            db.session.rollback()
        _set_auth_session(BOOTSTRAP_ADMIN_USERNAME, 'admin')
        return redirect(url_for('dashboard'))

    user, role = _pick_login_user(username)
    if not user or not check_password_hash(user.password_hash, password):
        flash("Invalid credentials")
        return _safe_render(ADMIN_LOGIN_HTML, allow_register=False), 401

    _set_auth_session(user.username, role)
    return _safe_login_redirect(role, request.args.get('next') or request.form.get('next'))

@app.route('/logout')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged', None)
    session.pop('admin_user', None)
    session.pop('police_logged', None)
    session.pop('gk_logged', None)
    session.pop('auth_role', None)
    session.pop('username', None)
    session.pop('user_id', None)
    return redirect(url_for('admin_login'))

@app.route('/register', methods=['GET', 'POST'])
@app.route('/admin/register', methods=['GET', 'POST'])
def admin_register():
    flash("Registration is disabled. Use the bootstrap login first, then add admins/police/GK from inside the app.")
    return redirect(url_for('admin_login'))

@app.route('/gk', methods=['GET', 'POST'])
@app.route('/GK', methods=['GET', 'POST'])
@app.route('/gk/login', methods=['GET', 'POST'])
@app.route('/GK/login', methods=['GET', 'POST'])
def gk_login():
    if request.method == 'GET':
        if _current_role() == 'gk':
            return redirect(url_for('gk_dashboard'))
        return _safe_render(GK_LOGIN_HTML, allow_register=False)

    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    if not username or not password:
        flash("Missing username or password")
        return _safe_render(GK_LOGIN_HTML, allow_register=False), 400

    user, role = _pick_login_user(username)
    if role != 'gk' or not user or not check_password_hash(user.password_hash, password):
        flash("Invalid GK credentials")
        return _safe_render(GK_LOGIN_HTML, allow_register=False), 401

    _set_auth_session(user.username, 'gk')
    return redirect(url_for('gk_dashboard'))

@app.route('/gk/logout')
@app.route('/GK/logout')
def gk_logout():
    session.clear()
    return redirect(url_for('gk_login'))

@app.route('/dashboard')

def dashboard():
    if _current_role() != 'admin':
        return redirect(url_for('admin_login'))
    return _safe_render(DASHBOARD_HTML)

# keep /friendly for compatibility
@app.route('/friendly')
def friendly():
    if _current_role() != 'admin':
        return redirect(url_for('admin_login'))
    return _safe_render(DASHBOARD_HTML)

GK_LOGIN_HTML = """
<!doctype html>
<html>
<head><meta charset="utf-8"><title>GK Login</title><meta name="viewport" content="width=device-width, initial-scale=1" />
<style>
body{font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,Arial;background:linear-gradient(180deg,#07111f 0%, #0b1220 100%);margin:0;padding:24px;color:#e5eefb;}
.card{max-width:560px;margin:0 auto;background:rgba(16,26,45,.96);border:1px solid rgba(148,163,184,.15);border-radius:24px;padding:24px;box-shadow:0 24px 64px rgba(0,0,0,.28);backdrop-filter: blur(12px);} 
input{width:100%;box-sizing:border-box;padding:12px;border:1px solid rgba(148,163,184,.18);border-radius:14px;margin-top:8px;background:#0b1324;color:#e5eefb;}
button,a{display:inline-block;margin-top:16px;padding:12px 16px;border-radius:999px;border:0;background:linear-gradient(135deg,#8b5cf6,#38bdf8);color:#fff;text-decoration:none;font-weight:800;cursor:pointer;}
.muted{color:#94a3b8;font-size:14px;line-height:1.5;}
.flash{background:rgba(239,68,68,.12);border:1px solid rgba(239,68,68,.25);color:#fecaca;padding:12px;border-radius:14px;margin-top:12px;}
</style>
</head>
<body>
  <div class="card">
    <h2>GK Login</h2>
    {% with messages = get_flashed_messages() %}
      {% if messages %}
        <div class="flash">{{ messages[0] }}</div>
      {% endif %}
    {% endwith %}
    <form method="post" action="{{ request.path }}">
      <label>Username</label>
      <input name="username" required>
      <label>Password</label>
      <input name="password" type="password" required>
      <button type="submit">Login</button>
    </form>
    <p class="muted">GK users are created from Users & Access. After login, the GK panel opens here.</p>
  </div>
</body>
</html>
"""

GK_DASHBOARD_HTML = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>GK Panel</title>
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>
    body{font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,Arial;background:linear-gradient(180deg,#07111f 0%, #0b1220 100%);margin:0;padding:24px;color:#e5eefb;}
    .wrap{max-width:1100px;margin:0 auto;}
    .card{background:#fff;color:#0f172a;border:1px solid #e2e8f0;border-radius:18px;padding:18px;box-shadow:0 10px 24px rgba(0,0,0,.05);margin-bottom:14px;}
    a,button{display:inline-block;margin-top:10px;padding:10px 14px;border-radius:10px;border:0;background:#0b84ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer;}
    input,textarea,select{width:100%;box-sizing:border-box;padding:12px;border:1px solid #e2e8f0;border-radius:10px;margin-top:8px;}
    table{width:100%;border-collapse:collapse;}
    td,th{padding:10px;border-bottom:1px solid #e2e8f0;text-align:left;vertical-align:top;}
    .grid{display:grid;grid-template-columns:1.2fr .8fr;gap:14px;}
    .muted{color:#64748b;font-size:14px;}
    .pill{display:inline-block;padding:6px 10px;border-radius:999px;background:#eff6ff;color:#1d4ed8;font-weight:700;margin-right:6px;margin-top:6px;}
    @media (max-width: 900px){ .grid{grid-template-columns:1fr;} }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div style="display:flex;justify-content:space-between;gap:10px;flex-wrap:wrap;align-items:center;">
        <div>
          <h1 style="margin:0;">GK Panel</h1>
          <div class="muted">Send a message to one user, overspeeders, all users, or a zone/road.</div>
        </div>
        <div>
          <a href="{{ url_for('gk_logout') }}">Logout</a>
          <a href="{{ url_for('admin_messages') }}">Open Messages Center</a>
        </div>
      </div>
    </div>

    <div class="grid">
      <div class="card">
        <h2>Quick send</h2>
        <form method="post" action="{{ url_for('admin_messages') }}" id="quickSendForm">
          <label>Title</label>
          <input name="title" value="Attention" required>
          <label>Message</label>
          <textarea name="body" rows="5" placeholder="Type your text here..." required>Hi</textarea>
          <input type="hidden" name="target_type" value="single">
          <label>Device ID</label>
          <input name="target_device_id" placeholder="Paste a device ID from search results">
          <button type="submit">Send to one user</button>
        </form>
      </div>
      <div class="card">
        <h2>Recent messages</h2>
        {% if recent_messages %}
          <table>
            <thead><tr><th>Title</th><th>Target</th><th>Recipients</th></tr></thead>
            <tbody>
            {% for m in recent_messages %}
              <tr>
                <td>{{ m.title }}</td>
                <td>{{ m.target_type }}</td>
                <td>{{ m.recipient_count }}</td>
              </tr>
            {% endfor %}
            </tbody>
          </table>
        {% else %}
          <div class="muted">No messages have been sent yet.</div>
        {% endif %}
        <p class="muted" style="margin-top:12px;">Messages sent here are also pushed to devices live through the app socket and the device message endpoint.</p>
      </div>
    </div>
  </div>
  <script>
  // The actual native pop-up delivery is done through the socket event and /device/messages on the app side.
  </script>
</body>
</html>
"""

@app.route('/gk/dashboard')
@app.route('/GK/dashboard')
def gk_dashboard():
    if _current_role() not in {'gk', 'admin'}:
        return redirect(url_for('gk_login'))
    recent_messages = BroadcastMessage.query.order_by(BroadcastMessage.created_at.desc()).limit(10).all()
    return _safe_render(GK_DASHBOARD_HTML, recent_messages=recent_messages)

MESSAGES_HTML = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Messages Center</title>
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>
    body{font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,Arial;background:linear-gradient(180deg,#07111f 0%, #0b1220 100%);margin:0;padding:24px;color:#e5eefb;}
    .wrap{max-width:1400px;margin:0 auto;}
    .card{background:#fff;color:#0f172a;border:1px solid #e2e8f0;border-radius:18px;padding:18px;box-shadow:0 10px 24px rgba(0,0,0,.05);margin-bottom:14px;}
    a,button{display:inline-block;margin-top:10px;padding:10px 14px;border-radius:10px;border:0;background:#0b84ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer;}
    input,textarea,select{width:100%;box-sizing:border-box;padding:12px;border:1px solid #e2e8f0;border-radius:10px;margin-top:8px;}
    table{width:100%;border-collapse:collapse;}
    td,th{padding:10px;border-bottom:1px solid #e2e8f0;text-align:left;vertical-align:top;}
    .grid{display:grid;grid-template-columns:1.1fr .9fr;gap:14px;}
    .mini-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;}
    .muted{color:#64748b;font-size:14px;}
    .pill{display:inline-block;padding:6px 10px;border-radius:999px;background:#eff6ff;color:#1d4ed8;font-weight:700;margin-right:6px;margin-top:6px;}
    .result{border:1px solid #e2e8f0;border-radius:12px;padding:10px;margin-top:10px;display:flex;justify-content:space-between;gap:10px;align-items:flex-start;}
    .result small{color:#64748b;display:block;}
    @media (max-width: 900px){ .grid,.mini-grid{grid-template-columns:1fr;} }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div style="display:flex;justify-content:space-between;gap:10px;flex-wrap:wrap;align-items:center;">
        <div>
          <h1 style="margin:0;">Messages Center</h1>
          <div class="muted">Send one message to one user, overspeeders, a zone, a road, a county, or everyone.</div>
        </div>
        <div>
          <a href="{{ url_for('dashboard') }}">Dashboard</a>
          <a href="{{ url_for('admin_admins') }}">Users</a>
          <a href="{{ url_for('admin_traffic') }}">Traffic</a>
          <a href="{{ url_for('admin_speeders') }}">Speeders</a>
          <a href="{{ url_for('admin_logout') }}">Logout</a>
        </div>
      </div>
      {% with messages = get_flashed_messages() %}
        {% if messages %}<div class="pill">{{ messages[0] }}</div>{% endif %}
      {% endwith %}
    </div>

    <div class="grid">
      <div class="card">
        <h2>Compose message</h2>
        <form method="post" action="{{ url_for('admin_messages') }}">
          <div class="mini-grid">
            <div>
              <label>Title</label>
              <input name="title" required placeholder="Enter title">
            </div>
            <div>
              <label>Target type</label>
              <select name="target_type" id="target_type" onchange="toggleTargetFields()">
                <option value="single">One user</option>
                <option value="all">All users</option>
                <option value="overspeeders">Overspeeders</option>
                <option value="zone">Zone / region</option>
                <option value="road">Road</option>
                <option value="county">County / text match</option>
                <option value="search">Search text</option>
              </select>
            </div>
          </div>

          <label>Message</label>
          <textarea name="body" rows="5" required placeholder="Type your message here"></textarea>

          <div id="field_single" style="margin-top:10px;">
            <label>Target device ID</label>
            <input name="target_device_id" id="target_device_id" placeholder="Use the search results on the right">
          </div>
          <div id="field_overspeeders" style="margin-top:10px;display:none;">
            <label>Overspeed threshold (km/h)</label>
            <input name="min_kmh" value="80">
          </div>
          <div id="field_zone" style="margin-top:10px;display:none;">
            <label>Zone</label>
            <select name="zone_id">
              <option value="">Select zone</option>
              {% for z in zones %}
                <option value="{{ z.id }}">{{ z.name }} ({{ z.scope }})</option>
              {% endfor %}
            </select>
          </div>
          <div id="field_road" style="margin-top:10px;display:none;">
            <label>Road</label>
            <select name="road_id">
              <option value="">Select road</option>
              {% for r in roads %}
                <option value="{{ r.id }}">{{ r.name }} — {{ r.speed_limit_kmh }} km/h</option>
              {% endfor %}
            </select>
          </div>
          <div id="field_county" style="margin-top:10px;display:none;">
            <label>County / text match</label>
            <input name="county" placeholder="e.g. Nairobi">
          </div>
          <div id="field_search" style="margin-top:10px;display:none;">
            <label>Search text</label>
            <input name="query" placeholder="plate, owner, make, model">
          </div>

          <button type="submit">Send message</button>
        </form>
      </div>

      <div class="card">
        <h2>Search users</h2>
        <input id="searchBox" placeholder="Search plate, owner, make, model..." oninput="searchDevices()">
        <div class="muted" style="margin-top:8px;">Tap a result to copy its device ID into the message form.</div>
        <div id="searchResults" style="margin-top:10px;max-height:520px;overflow:auto;"></div>
      </div>
    </div>

    <div class="grid">
      <div class="card">
        <h2>Recent messages</h2>
        <table>
          <thead><tr><th>Title</th><th>Target</th><th>Recipients</th><th>Created</th></tr></thead>
          <tbody>
          {% for m in recent_messages %}
            <tr>
              <td>{{ m.title }}</td>
              <td>{{ m.target_type }} {% if m.target_value %}({{ m.target_value }}){% endif %}</td>
              <td>{{ m.recipient_count }}</td>
              <td>{{ m.created_at.isoformat() if m.created_at else '' }}</td>
            </tr>
          {% endfor %}
          </tbody>
        </table>
      </div>
      <div class="card">
        <h2>Quick targets</h2>
        <div class="muted">One-click helper buttons</div>
        <div style="margin-top:10px;display:flex;flex-wrap:wrap;gap:8px;">
          <button type="button" onclick="setTarget('all')">All users</button>
          <button type="button" onclick="setTarget('overspeeders')">Overspeeders</button>
          <button type="button" onclick="setTarget('zone')">Zone</button>
          <button type="button" onclick="setTarget('road')">Road</button>
          <button type="button" onclick="setTarget('county')">County</button>
          <button type="button" onclick="setTarget('single')">One user</button>
        </div>
        <div class="muted" style="margin-top:12px;">The app can pop these up through the native message listener or the live socket event.</div>
      </div>
    </div>
  </div>

<script>
function toggleTargetFields(){
  const t = document.getElementById('target_type').value;
  const ids = ['field_single','field_overspeeders','field_zone','field_road','field_county','field_search'];
  ids.forEach(id => document.getElementById(id).style.display = 'none');
  if (t === 'single') document.getElementById('field_single').style.display = 'block';
  if (t === 'overspeeders') document.getElementById('field_overspeeders').style.display = 'block';
  if (t === 'zone') document.getElementById('field_zone').style.display = 'block';
  if (t === 'road') document.getElementById('field_road').style.display = 'block';
  if (t === 'county') document.getElementById('field_county').style.display = 'block';
  if (t === 'search') document.getElementById('field_search').style.display = 'block';
}
function setTarget(v){
  document.getElementById('target_type').value = v;
  toggleTargetFields();
}
async function searchDevices(){
  const q = document.getElementById('searchBox').value.trim();
  const res = await fetch('/admin/message/search-devices?q=' + encodeURIComponent(q), {cache:'no-store'});
  const j = await res.json();
  const box = document.getElementById('searchResults');
  const devices = j.devices || [];
  if (!devices.length){ box.innerHTML = '<div class="muted" style="padding:10px;">No devices found.</div>'; return; }
  box.innerHTML = devices.map(d => {
    const line1 = [d.owner, d.car_name, d.car_model, d.plate].filter(Boolean).join(' • ');
    const line2 = [d.speed_kmh ? d.speed_kmh + ' km/h' : '', d.ts || ''].filter(Boolean).join(' • ');
    return `<div class="result"><div><strong>${line1 || d.id}</strong><small>${line2}</small><small>${d.id}</small></div><div><button type="button" onclick="pickDevice('${d.id}')">Use</button></div></div>`;
  }).join('');
}
function pickDevice(id){
  document.getElementById('target_type').value = 'single';
  toggleTargetFields();
  document.getElementById('target_device_id').value = id;
  window.scrollTo({top:0, behavior:'smooth'});
}
toggleTargetFields();
searchDevices();
</script>
</body>
</html>
"""
@app.route('/admin/devices')
def admin_devices():
    if _current_role() != 'admin':
        return redirect(url_for('admin_login'))
    devices = Device.query.all()
    out = []
    prune_active_devices()
    for d in devices:
        snap = Snapshot.query.filter_by(device_id=d.id).order_by(Snapshot.ts.desc()).first()
        last = None
        if snap:
            parsed_raw = None
            if snap.raw:
                try:
                    parsed_raw = json.loads(snap.raw)
                except Exception:
                    parsed_raw = snap.raw
            last = {
                "ts": snap.ts.isoformat(),
                "lat": snap.lat,
                "lon": snap.lon,
                "speed_mps": round(snap.speed_mps, 3),
                "bearing_deg": round(snap.bearing_deg, 1),
                "raw": parsed_raw
            }
        else:
            with active_devices_lock:
                entry = active_devices.get(d.id)
            if entry:
                parsed_raw = entry.get("raw")
                last = {
                    "ts": entry.get("ts").isoformat() if entry.get("ts") else None,
                    "lat": entry.get("lat"),
                    "lon": entry.get("lon"),
                    "speed_mps": round(entry.get("speed_mps") or 0.0, 3),
                    "bearing_deg": round(entry.get("bearing_deg") or 0.0, 1),
                    "raw": parsed_raw
                }

        out.append({
            "id": d.id,
            "owner": d.owner,
            "car_name": d.car_name,
            "car_model": d.car_model,
            "plate": d.plate,
            "extra": d.extra,
            "created_at": d.created_at.isoformat() if d.created_at else None,
            "revoked": bool(d.revoked),
            "last_snapshot": last,
            "connected": bool(connected_sockets.get(d.id))
        })
    return jsonify({"devices": out})

@app.route('/admin/device/<device_id>/json')
def admin_device_json(device_id):
    d = Device.query.get_or_404(device_id)
    snaps = _recent_snapshots_for_device(device_id, limit=20)
    snaps_out = []
    for s in snaps:
        parsed_raw = None
        if s.raw:
            try:
                parsed_raw = json.loads(s.raw)
            except Exception:
                parsed_raw = s.raw
        snaps_out.append({
            "ts": s.ts.isoformat(),
            "lat": s.lat,
            "lon": s.lon,
            "speed_mps": s.speed_mps,
            "bearing_deg": s.bearing_deg,
            "source": s.source,
            "raw": parsed_raw
        })

    if not snaps_out:
        with active_devices_lock:
            entry = active_devices.get(device_id)
        if entry:
            snaps_out.append({
                "ts": entry.get("ts").isoformat() if entry.get("ts") else None,
                "lat": entry.get("lat"),
                "lon": entry.get("lon"),
                "speed_mps": entry.get("speed_mps"),
                "bearing_deg": entry.get("bearing_deg"),
                "source": entry.get("source", "app"),
                "raw": entry.get("raw")
            })

    last = snaps_out[0] if snaps_out else None
    return jsonify({
        "device": {
            "id": d.id,
            "owner": d.owner,
            "car_name": d.car_name,
            "car_model": d.car_model,
            "plate": d.plate,
            "extra": d.extra,
            "created_at": d.created_at.isoformat() if d.created_at else None,
            "revoked": bool(d.revoked)
        },
        "last_snapshot": last,
        "snapshots": snaps_out,
        "connected": bool(connected_sockets.get(d.id))
    })

@app.route('/admin/device/<device_id>/revoke', methods=['POST'])
def admin_device_revoke(device_id):
    d = Device.query.get_or_404(device_id)
    d.revoked = True
    db.session.commit()
    sids = connected_sockets.pop(device_id, None)
    with active_devices_lock:
        active_devices.pop(device_id, None)
    return jsonify({"ok": True, "revoked": True})

# -------------------------
# Messaging helpers / routes
# -------------------------

def _latest_snapshots_map():
    latest = {}
    for snap in Snapshot.query.order_by(Snapshot.ts.desc()).all():
        if snap.device_id not in latest:
            latest[snap.device_id] = snap
    return latest


def _device_blob(device):
    extra_obj = None
    if device.extra:
        try:
            extra_obj = json.loads(device.extra)
        except Exception:
            extra_obj = device.extra
    return {
        'id': device.id,
        'token': device.token,
        'owner': device.owner,
        'car_name': device.car_name,
        'car_model': device.car_model,
        'plate': device.plate,
        'extra': extra_obj,
    }


def _device_text_match(device, q):
    if not q:
        return True
    q = q.lower().strip()
    hay = ' '.join([
        str(device.id or ''), str(device.owner or ''), str(device.car_name or ''),
        str(device.car_model or ''), str(device.plate or ''), str(device.extra or '')
    ]).lower()
    return q in hay


def _target_devices_for_message(body):
    latest = _latest_snapshots_map()
    target_type = (body.get('target_type') or 'single').strip().lower()
    target_value = (body.get('target_value') or '').strip()
    recipients = []

    devices = Device.query.filter_by(revoked=False).all()
    if target_type == 'all':
        recipients = devices
    elif target_type == 'single':
        did = (body.get('target_device_id') or target_value or '').strip()
        if did:
            d = Device.query.filter_by(id=did, revoked=False).first()
            recipients = [d] if d else []
    elif target_type == 'overspeeders':
        min_kmh = float(body.get('min_kmh') or 80.0)
        recipients = []
        for d in devices:
            snap = latest.get(d.id)
            if snap and (float(snap.speed_mps or 0.0) * 3.6) >= min_kmh:
                recipients.append(d)
    elif target_type == 'road':
        road_id = (body.get('road_id') or target_value or '').strip()
        road = Road.query.get(road_id) if road_id else None
        if road and road.center_lat is not None and road.center_lon is not None and road.radius_m is not None:
            for d in devices:
                snap = latest.get(d.id)
                if snap and haversine_m(snap.lat, snap.lon, road.center_lat, road.center_lon) <= float(road.radius_m):
                    recipients.append(d)
    elif target_type == 'zone':
        zone_id = (body.get('zone_id') or target_value or '').strip()
        zone = TrafficZone.query.get(zone_id) if zone_id else None
        if zone:
            for d in devices:
                snap = latest.get(d.id)
                if snap and _zone_matches(zone, snap.lat, snap.lon):
                    recipients.append(d)
    elif target_type == 'county':
        county = (body.get('county') or target_value or '').strip().lower()
        if county:
            for d in devices:
                if county in ' '.join([str(d.owner or ''), str(d.car_name or ''), str(d.car_model or ''), str(d.plate or ''), str(d.extra or '')]).lower():
                    recipients.append(d)
    elif target_type == 'search':
        q = (body.get('query') or target_value or '').strip()
        recipients = [d for d in devices if _device_text_match(d, q)]
    else:
        recipients = []

    seen = set()
    out = []
    for d in recipients:
        if not d or d.id in seen:
            continue
        seen.add(d.id)
        out.append(d)
    return out


def _serialize_message(msg, recipient_count=None):
    return {
        'id': msg.id,
        'title': msg.title,
        'body': msg.body,
        'target_type': msg.target_type,
        'target_value': msg.target_value,
        'creator_role': msg.creator_role,
        'creator_username': msg.creator_username,
        'recipient_count': recipient_count if recipient_count is not None else msg.recipient_count,
        'created_at': msg.created_at.isoformat() if msg.created_at else None,
    }


def _pending_messages_for_device(device_id, mark_delivered=True):
    rows = (
        db.session.query(BroadcastDelivery, BroadcastMessage)
        .join(BroadcastMessage, BroadcastMessage.id == BroadcastDelivery.message_id)
        .filter(BroadcastDelivery.device_id == device_id)
        .filter(BroadcastDelivery.delivered_at.is_(None))
        .order_by(BroadcastMessage.created_at.asc())
        .all()
    )
    messages = []
    now = datetime.utcnow()
    for delivery, msg in rows:
        messages.append(_serialize_message(msg))
        if mark_delivered:
            delivery.delivered_at = now
    if mark_delivered and rows:
        db.session.commit()
    return messages


def _create_message_and_dispatch(body):
    title = (body.get('title') or '').strip()
    message_body = (body.get('body') or body.get('message') or '').strip()
    if not title or not message_body:
        return None, 'title and body are required'

    recipients = _target_devices_for_message(body)
    msg = BroadcastMessage(
        title=title,
        body=message_body,
        target_type=(body.get('target_type') or 'single').strip().lower(),
        target_value=(body.get('target_value') or body.get('target_device_id') or body.get('zone_id') or body.get('road_id') or body.get('county') or body.get('query') or ''),
        creator_role=_current_role() or 'admin',
        creator_username=session.get('username') or session.get('admin_user') or '' ,
        recipient_count=len(recipients),
    )
    db.session.add(msg)
    db.session.commit()

    now = datetime.utcnow()
    for d in recipients:
        db.session.add(BroadcastDelivery(message_id=msg.id, device_id=d.id, delivered_at=None, read_at=None))
    db.session.commit()

    payload = _serialize_message(msg, recipient_count=len(recipients))
    payload['device_ids'] = [d.id for d in recipients]
    payload['app_action'] = 'popup'

    for d in recipients:
        try:
            send_ws_to_device(d.id, 'admin_message', payload)
        except Exception:
            pass
    return payload, None


@app.route('/admin/messages', methods=['GET', 'POST'])
def admin_messages():
    role = _current_role()
    if role not in {'admin', 'gk'}:
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        body = request.form.to_dict(flat=True)
        if not body.get('title') and request.is_json:
            body = request.get_json(force=True, silent=True) or {}
        payload, err = _create_message_and_dispatch(body)
        if err:
            flash(err)
        else:
            flash(f"Message sent to {payload.get('recipient_count', 0)} device(s)")
        if request.is_json:
            if err:
                return jsonify({'error': err}), 400
            return jsonify({'ok': True, 'message': payload})
        return redirect(url_for('admin_messages'))

    latest = _latest_snapshots_map()
    devices = Device.query.filter_by(revoked=False).order_by(Device.created_at.desc()).all()
    speeders = []
    for d in devices:
        snap = latest.get(d.id)
        if snap and (float(snap.speed_mps or 0.0) * 3.6) >= 80:
            speeders.append(d)
    roads = Road.query.order_by(Road.created_at.desc()).all()
    zones = TrafficZone.query.order_by(TrafficZone.created_at.desc()).all()
    recent_messages = BroadcastMessage.query.order_by(BroadcastMessage.created_at.desc()).limit(20).all()
    return _safe_render(MESSAGES_HTML, devices=devices, speeders=speeders, roads=roads, zones=zones, recent_messages=recent_messages)


@app.route('/admin/message/search-devices')
def admin_message_search_devices():
    role = _current_role()
    if role not in {'admin', 'gk'}:
        return redirect(url_for('admin_login'))
    q = (request.args.get('q') or '').strip().lower()
    devices = Device.query.filter_by(revoked=False).order_by(Device.created_at.desc()).all()
    latest = _latest_snapshots_map()
    out = []
    for d in devices:
        if q and q not in ' '.join([str(d.id or ''), str(d.owner or ''), str(d.car_name or ''), str(d.car_model or ''), str(d.plate or ''), str(d.extra or '')]).lower():
            continue
        snap = latest.get(d.id)
        out.append({
            'id': d.id,
            'owner': d.owner,
            'car_name': d.car_name,
            'car_model': d.car_model,
            'plate': d.plate,
            'extra': d.extra,
            'speed_kmh': round(float(snap.speed_mps or 0.0) * 3.6, 1) if snap else None,
            'ts': snap.ts.isoformat() if snap and snap.ts else None,
        })
        if len(out) >= 50:
            break
    return jsonify({'devices': out})


@app.route('/device/messages', methods=['GET'])
def device_messages():
    body = request.get_json(silent=True) or {}
    device = find_or_restore_from_request(body)
    if not device:
        try:
            device = require_auth_token()
        except Exception:
            return jsonify({'error': 'Missing or invalid token'}), 401
    messages = _pending_messages_for_device(device.id, mark_delivered=True)
    return jsonify({'ok': True, 'messages': messages})


@app.route('/admin/message/send', methods=['POST'])
def admin_message_send():
    role = _current_role()
    if role not in {'admin', 'gk'}:
        return redirect(url_for('admin_login'))
    body = request.get_json(force=True, silent=True) or request.form.to_dict(flat=True)
    payload, err = _create_message_and_dispatch(body)
    if err:
        return jsonify({'error': err}), 400
    return jsonify({'ok': True, 'message': payload})


@app.route('/admin/messages/html-snippet')
def admin_messages_html_snippet():
    role = _current_role()
    if role not in {'admin', 'gk'}:
        return redirect(url_for('admin_login'))
    return _safe_render(MESSAGES_HTML, devices=[], speeders=[], roads=[], zones=[], recent_messages=[])


@app.route('/admin/roads', methods=['GET', 'POST'])
def admin_roads():
    # GET: list roads
    if request.method == 'GET':
        # public to admin session or API token
        require_admin_api()
        roads = Road.query.order_by(Road.created_at.desc()).all()
        out = []
        for r in roads:
            out.append({
                "id": r.id,
                "name": r.name,
                "speed_limit_kmh": float(r.speed_limit_kmh),
                "center_lat": float(r.center_lat) if r.center_lat is not None else None,
                "center_lon": float(r.center_lon) if r.center_lon is not None else None,
                "radius_m": float(r.radius_m) if r.radius_m is not None else None,
                "created_at": r.created_at.isoformat() if r.created_at else None
            })
        return jsonify({"roads": out})

    # POST: create a road
    require_admin_api()
    body = request.get_json(force=True, silent=True) or {}
    name = body.get("name")
    speed_limit_kmh = body.get("speed_limit_kmh")
    center_lat = body.get("center_lat")
    center_lon = body.get("center_lon")
    radius_m = body.get("radius_m", 50.0)
    if not name or speed_limit_kmh is None:
        return jsonify({"error": "name and speed_limit_kmh required"}), 400
    try:
        r = Road(name=name, speed_limit_kmh=float(speed_limit_kmh))
        if center_lat is not None and center_lon is not None:
            r.center_lat = float(center_lat)
            r.center_lon = float(center_lon)
        r.radius_m = float(radius_m)
        db.session.add(r)
        db.session.commit()
        return jsonify({"ok": True, "road": {
            "id": r.id, "name": r.name, "speed_limit_kmh": r.speed_limit_kmh,
            "center_lat": r.center_lat, "center_lon": r.center_lon, "radius_m": r.radius_m
        }})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/admin/road/<road_id>', methods=['GET', 'DELETE'])
def admin_road_detail(road_id):
    require_admin_api()
    r = Road.query.get_or_404(road_id)
    if request.method == 'GET':
        return jsonify({
            "id": r.id, "name": r.name, "speed_limit_kmh": float(r.speed_limit_kmh),
            "center_lat": float(r.center_lat) if r.center_lat is not None else None,
            "center_lon": float(r.center_lon) if r.center_lon is not None else None,
            "radius_m": float(r.radius_m) if r.radius_m is not None else None,
            "created_at": r.created_at.isoformat() if r.created_at else None
        })
    # DELETE
    try:
        db.session.delete(r)
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/admin/overspeeds', methods=['GET'])
def admin_list_overspeeds():
    require_admin_api()
    limit = int(request.args.get("limit", 1000))
    query = OverspeedEvent.query.order_by(OverspeedEvent.ts.desc()).limit(limit).all()
    out = []
    for e in query:
        out.append({
            "id": e.id,
            "device_id": e.device_id,
            "road_id": e.road_id,
            "snapshot_id": e.snapshot_id,
            "ts": e.ts.isoformat() if e.ts else None,
            "speed_kmh": float(e.speed_kmh) if e.speed_kmh is not None else None,
            "lat": e.lat, "lon": e.lon,
            "raw": (json.loads(e.raw) if e.raw else None) if isinstance(e.raw, str) else e.raw
        })
    return jsonify({"overspeeds": out})

@app.route('/admin/speeders', methods=['GET'])
def admin_speeders():
    if _current_role() != 'admin':
        return redirect(url_for('admin_login'))
    min_kmh = float(request.args.get('min_kmh', 80))
    zone_id = (request.args.get('zone_id') or '').strip()
    zone = TrafficZone.query.filter_by(id=zone_id).first() if zone_id else None
    speeders = _speeders_snapshot(min_speed_kmh=min_kmh, zone=zone)
    zones = TrafficZone.query.order_by(TrafficZone.created_at.desc()).all()
    return _safe_render("""
<!doctype html>
<html>
<head>
  <meta charset='utf-8'>
  <title>Speeders</title>
  <meta name='viewport' content='width=device-width, initial-scale=1' />
  <style>
    body{font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,Arial;background:linear-gradient(180deg,#07111f 0%, #0b1220 100%);margin:0;padding:24px;color:#e5eefb;}
    .wrap{max-width:1100px;margin:0 auto;}
    .card{background:#fff;border:1px solid #e2e8f0;border-radius:14px;padding:18px;box-shadow:0 10px 24px rgba(0,0,0,.05);margin-bottom:14px;}
    table{width:100%;border-collapse:collapse;}
    td,th{padding:10px;border-bottom:1px solid #e2e8f0;text-align:left;}
    input,select{width:100%;box-sizing:border-box;padding:12px;border:1px solid #e2e8f0;border-radius:10px;margin-top:8px;}
    button,a{display:inline-block;margin-top:14px;padding:12px 16px;border-radius:10px;border:0;background:#0b84ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer;}
    .muted{color:#94a3b8;font-size:14px;line-height:1.5;}
    .pill{display:inline-block;padding:6px 10px;border-radius:999px;background:#eff6ff;color:#1d4ed8;font-weight:700;margin-right:6px;margin-top:6px;}
  </style>
</head>
<body>
  <div class='wrap'>
    <div class='card'>
      <div style='display:flex;justify-content:space-between;gap:10px;align-items:center;flex-wrap:wrap;'>
        <div>
          <h1 style='margin:0;'>Speeders</h1>
          <div class='muted'>Vehicles whose latest speed is at or above the chosen threshold.</div>
        </div>
        <div>
          <a href='{{ url_for("dashboard") }}'>Dashboard</a>
          <a href='{{ url_for("all_vehicles") }}'>All Vehicles</a>
          <a href='{{ url_for("admin_traffic") }}'>Traffic Search</a>
          <a href='{{ url_for("admin_logout") }}'>Logout</a>
        </div>
      </div>
    </div>

    <div class='card'>
      <form method='get'>
        <label>Minimum speed (km/h)</label>
        <input name='min_kmh' value='{{ min_kmh }}' />
        <label>Saved zone (optional)</label>
        <select name='zone_id'>
          <option value=''>-- any area --</option>
          {% for z in zones %}
            <option value='{{ z.id }}' {% if zone and zone.id == z.id %}selected{% endif %}>{{ z.name }} ({{ z.scope }})</option>
          {% endfor %}
        </select>
        <button type='submit'>Show speeders</button>
      </form>
    </div>

    <div class='card'>
      <h2 style='margin-top:0;'>{{ speeders|length }} vehicle(s)</h2>
      <table>
        <thead><tr><th>Plate</th><th>Owner</th><th>Vehicle</th><th>Speed</th><th>Last seen</th><th>Last location</th></tr></thead>
        <tbody>
        {% for v in speeders %}
          <tr>
            <td>{{ v.plate or '—' }}</td>
            <td>{{ v.owner or '—' }}</td>
            <td>{{ v.car_name or v.car_model or v.device_id }}</td>
            <td>{{ '%.1f km/h'|format(v.speed_kmh) }}</td>
            <td>{{ v.last_snapshot.ts or '—' }}</td>
            <td>{{ v.last_snapshot.lat if v.last_snapshot.lat is not none else '—' }}, {{ v.last_snapshot.lon if v.last_snapshot.lon is not none else '—' }}</td>
          </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
</body>
</html>
""", speeders=speeders, min_kmh=min_kmh, zone=zone, zones=zones)

# -------------------------
# Report generation utils (Excel & PDF)
# -------------------------

def generate_all_excel_bytes():
    """Create a simple workbook with devices, snapshots, roads, overspeeds."""
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel report generation")
    wb = Workbook()

    ws = wb.active
    ws.title = "devices"
    ws.append(["id", "owner", "car_name", "car_model", "plate", "created_at", "revoked"])
    for d in Device.query.all():
        ws.append([d.id, d.owner, d.car_name, d.car_model, d.plate,
                   d.created_at.isoformat() if d.created_at else "", bool(d.revoked)])

    ws = wb.create_sheet("snapshots")
    ws.append(["id", "device_id", "ts", "lat", "lon", "speed_mps", "speed_kmh", "bearing_deg", "source", "raw"])
    for s in Snapshot.query.order_by(Snapshot.ts.desc()).limit(2000).all():
        raw = ""
        if s.raw:
            try:
                raw = json.dumps(json.loads(s.raw))
            except Exception:
                raw = str(s.raw)
        ws.append([s.id, s.device_id, s.ts.isoformat() if s.ts else "", s.lat, s.lon, s.speed_mps,
                   round((s.speed_mps or 0.0) * 3.6, 2), s.bearing_deg, s.source, raw])

    ws = wb.create_sheet("roads")
    ws.append(["id", "name", "speed_limit_kmh", "center_lat", "center_lon", "radius_m", "created_at"])
    for r in Road.query.order_by(Road.created_at.desc()).all():
        ws.append([r.id, r.name, r.speed_limit_kmh, r.center_lat, r.center_lon, r.radius_m,
                   r.created_at.isoformat() if r.created_at else ""])

    ws = wb.create_sheet("overspeeds")
    ws.append(["id", "device_id", "road_id", "snapshot_id", "ts", "speed_kmh", "lat", "lon", "raw"])
    for o in OverspeedEvent.query.order_by(OverspeedEvent.ts.desc()).limit(2000).all():
        raw = ""
        if o.raw:
            try:
                raw = json.dumps(json.loads(o.raw))
            except Exception:
                raw = str(o.raw)
        ws.append([o.id, o.device_id, o.road_id, o.snapshot_id, o.ts.isoformat() if o.ts else "",
                   o.speed_kmh, o.lat, o.lon, raw])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_road_excel_bytes(road_id):
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel report generation")
    wb = Workbook()
    road = Road.query.get_or_404(road_id)

    ws = wb.active
    ws.title = "road_info"
    ws.append(["id", "name", "speed_limit_kmh", "center_lat", "center_lon", "radius_m", "created_at"])
    ws.append([road.id, road.name, road.speed_limit_kmh, road.center_lat, road.center_lon, road.radius_m,
               road.created_at.isoformat() if road.created_at else ""])

    ws = wb.create_sheet("overspeeds")
    ws.append(["id", "device_id", "snapshot_id", "ts", "speed_kmh", "lat", "lon", "raw"])
    for o in OverspeedEvent.query.filter_by(road_id=road_id).order_by(OverspeedEvent.ts.desc()).all():
        raw = ""
        if o.raw:
            try:
                raw = json.dumps(json.loads(o.raw))
            except Exception:
                raw = str(o.raw)
        ws.append([o.id, o.device_id, o.snapshot_id, o.ts.isoformat() if o.ts else "",
                   o.speed_kmh, o.lat, o.lon, raw])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


    if pd is None:
        raise RuntimeError("pandas is required for Excel report generation")
    r = Road.query.get_or_404(road_id)
    # basic road info sheet + overspeeds sheet
    overs = OverspeedEvent.query.filter_by(road_id=road_id).order_by(OverspeedEvent.ts.desc()).all()
    overs_rows = []
    for o in overs:
        parsed = None
        if o.raw:
            try:
                parsed = json.loads(o.raw)
            except Exception:
                parsed = o.raw
        overs_rows.append({
            "id": o.id, "device_id": o.device_id, "snapshot_id": o.snapshot_id,
            "ts": o.ts.isoformat() if o.ts else None,
            "speed_kmh": o.speed_kmh, "lat": o.lat, "lon": o.lon, "raw": json.dumps(parsed) if parsed is not None else None
        })
    df_overs = pd.DataFrame(overs_rows)
    df_info = pd.DataFrame([{
        "id": r.id, "name": r.name, "speed_limit_kmh": r.speed_limit_kmh,
        "center_lat": r.center_lat, "center_lon": r.center_lon, "radius_m": r.radius_m, "created_at": r.created_at.isoformat() if r.created_at else None
    }])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_info.to_excel(writer, sheet_name='road_info', index=False)
        df_overs.to_excel(writer, sheet_name='overspeeds', index=False)
        writer.save()
    output.seek(0)
    return output.getvalue()

def generate_all_pdf_bytes():
    """
    Attempt to create a simple PDF summary using reportlab.
    If reportlab not available, raise an error to let caller return a helpful message.
    """
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab library is required for PDF generation")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph("Beacon — Full App Report", styles['Heading1']))
    elems.append(Spacer(1, 12))

    # devices table (brief)
    devs = Device.query.all()
    dev_table_data = [["id", "owner", "car_name", "plate", "created_at", "revoked"]]
    for d in devs:
        dev_table_data.append([d.id, d.owner or "", d.car_name or "", d.plate or "", d.created_at.isoformat() if d.created_at else "", str(bool(d.revoked))])
    t = Table(dev_table_data, repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.lightblue),('GRID',(0,0),(-1,-1),0.25,colors.grey)]))
    elems.append(Paragraph("Devices", styles['Heading2']))
    elems.append(t)
    elems.append(Spacer(1,12))

    # roads
    roads = Road.query.all()
    road_table = [["id","name","speed_limit_kmh","center_lat","center_lon","radius_m"]]
    for r in roads:
        road_table.append([r.id, r.name, str(r.speed_limit_kmh), str(r.center_lat or ""), str(r.center_lon or ""), str(r.radius_m or "")])
    elems.append(Paragraph("Roads", styles['Heading2']))
    tr = Table(road_table, repeatRows=1)
    tr.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.lightblue),('GRID',(0,0),(-1,-1),0.25,colors.grey)]))
    elems.append(tr)
    elems.append(Spacer(1,12))

    # overspeeds (recent N)
    overs = OverspeedEvent.query.order_by(OverspeedEvent.ts.desc()).limit(500).all()
    overs_table = [["ts","device_id","road_id","speed_kmh","lat","lon"]]
    for o in overs:
        overs_table.append([o.ts.isoformat() if o.ts else "", o.device_id, o.road_id, str(o.speed_kmh), str(o.lat), str(o.lon)])
    elems.append(Paragraph("Recent Overspeed Events (most recent 500)", styles['Heading2']))
    to = Table(overs_table, repeatRows=1)
    to.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.lightblue),('GRID',(0,0),(-1,-1),0.25,colors.grey)]))
    elems.append(to)

    doc.build(elems)
    buffer.seek(0)
    return buffer.getvalue()

def generate_road_pdf_bytes(road_id):
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab library is required for PDF generation")
    r = Road.query.get_or_404(road_id)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph(f"Road Report — {r.name}", styles['Heading1']))
    elems.append(Paragraph(f"Speed limit: {r.speed_limit_kmh} km/h — Radius: {r.radius_m} m", styles['Normal']))
    elems.append(Spacer(1,12))

    overs = OverspeedEvent.query.filter_by(road_id=road_id).order_by(OverspeedEvent.ts.desc()).limit(1000).all()
    tdata = [["ts","device_id","speed_kmh","lat","lon"]]
    for o in overs:
        tdata.append([o.ts.isoformat() if o.ts else "", o.device_id, str(o.speed_kmh), str(o.lat), str(o.lon)])
    t = Table(tdata, repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.lightblue),('GRID',(0,0),(-1,-1),0.25,colors.grey)]))
    elems.append(t)

    doc.build(elems)
    buffer.seek(0)
    return buffer.getvalue()

# -------------------------
# Report endpoints
# -------------------------
@app.route('/report/all.xlsx')
def report_all_xlsx():
    try:
        require_admin_api()
    except Exception:
        # allow admin UI access via session cookie, else require admin token
        return abort(401, "Admin access required")
    try:
        data = generate_all_excel_bytes()
        return send_file(io.BytesIO(data), download_name="beacon_full_report.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/report/road/<road_id>.xlsx')
def report_road_xlsx(road_id):
    try:
        require_admin_api()
    except Exception:
        return abort(401, "Admin access required")
    try:
        data = generate_road_excel_bytes(road_id)
        road = Road.query.get(road_id)
        name_safe = (road.name[:30].replace(" ", "_") if road else road_id)
        return send_file(io.BytesIO(data), download_name=f"road_{name_safe}_report.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/report/all.pdf')
def report_all_pdf():
    try:
        require_admin_api()
    except Exception:
        return abort(401, "Admin access required")
    try:
        data = generate_all_pdf_bytes()
        return send_file(io.BytesIO(data), download_name="beacon_full_report.pdf", as_attachment=True, mimetype="application/pdf")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/report/road/<road_id>.pdf')
def report_road_pdf(road_id):
    try:
        require_admin_api()
    except Exception:
        return abort(401, "Admin access required")
    try:
        data = generate_road_pdf_bytes(road_id)
        road = Road.query.get(road_id)
        name_safe = (road.name[:30].replace(" ", "_") if road else road_id)
        return send_file(io.BytesIO(data), download_name=f"road_{name_safe}_report.pdf", as_attachment=True, mimetype="application/pdf")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# -------------------------
# Police / Watch features (drop-in)
# -------------------------
# Plate sightings model (camera / manual plate ingestion)
class PlateSighting(db.Model):
    id = db.Column(db.String(36), primary_key=True, default=lambda: uuid.uuid4().hex)
    plate = db.Column(db.String(64), nullable=False, index=True)
    lat = db.Column(db.Float, nullable=True)
    lon = db.Column(db.Float, nullable=True)
    ts = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    source = db.Column(db.String(64), default="camera")  # e.g., camera, manual
    raw = db.Column(db.Text, nullable=True)

# Admin: watchlist management (plates)
@app.route('/admin/watchlist', methods=['GET', 'POST'])
def admin_watchlist():
    require_admin_api()
    if request.method == 'GET':
        q = Watchlist.query.order_by(Watchlist.created_at.desc()).all()
        out = [{"id": w.id, "plate": w.plate, "label": w.label, "created_at": w.created_at.isoformat()} for w in q]
        return jsonify({"watchlist": out})
    body = request.get_json(force=True, silent=True) or {}
    plate = (body.get("plate") or "").strip()
    label = body.get("label")
    if not plate:
        return jsonify({"error": "plate required"}), 400
    try:
        w = Watchlist(plate=plate.upper(), label=label)
        db.session.add(w)
        db.session.commit()
        return jsonify({"ok": True, "watch": {"id": w.id, "plate": w.plate, "label": w.label}})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/admin/watchlist/<watch_id>', methods=['DELETE'])
def admin_watchlist_delete(watch_id):
    require_admin_api()
    w = Watchlist.query.get_or_404(watch_id)
    try:
        db.session.delete(w)
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# Police/watch API: query recent overspeed events (with device owner/vehicle meta)
@app.route('/watch/overspeeds', methods=['GET'])
def watch_overspeeds():
    require_admin_api()
    since = request.args.get("since")  # ISO timestamp (optional)
    road_id = request.args.get("road_id")
    plate = request.args.get("plate")
    limit = int(request.args.get("limit", 200))

    q = OverspeedEvent.query
    if since:
        try:
            dt = datetime.fromisoformat(since)
            q = q.filter(OverspeedEvent.ts >= dt)
        except Exception:
            pass
    if road_id:
        q = q.filter(OverspeedEvent.road_id == road_id)
    if plate:
        # find any device ids with that plate (case-insensitive)
        devs = Device.query.filter(db.func.upper(Device.plate) == plate.strip().upper()).all()
        dev_ids = [d.id for d in devs] if devs else []
        if dev_ids:
            q = q.filter(OverspeedEvent.device_id.in_(dev_ids))
        else:
            return jsonify({"overspeeds": []})

    q = q.order_by(OverspeedEvent.ts.desc()).limit(limit).all()
    out = []
    for e in q:
        dev_meta = {}
        try:
            d = Device.query.get(e.device_id)
            if d:
                dev_meta = {"owner": d.owner, "car_name": d.car_name, "car_model": d.car_model, "plate": d.plate}
        except Exception:
            dev_meta = {}
        out.append({
            "id": e.id,
            "device_id": e.device_id,
            "road_id": e.road_id,
            "ts": e.ts.isoformat() if e.ts else None,
            "speed_kmh": float(e.speed_kmh) if e.speed_kmh is not None else None,
            "lat": e.lat,
            "lon": e.lon,
            "device": dev_meta,
            "raw": (json.loads(e.raw) if e.raw and isinstance(e.raw, str) else e.raw)
        })
    return jsonify({"overspeeds": out})

# Plate ingestion endpoint (cameras, manual entries)
# Example body: { "plate": "KAA123A", "lat": -1.29, "lon": 36.82, "source": "camera-lpr-1", "raw": {...} }
@app.route('/ingest/plate', methods=['POST'])
def ingest_plate():
    # allow camera systems (no device token). Protect by admin token if desired:
    # require_admin_api()  <-- comment/uncomment depending on desired security
    body = request.get_json(force=True, silent=True) or {}
    plate = (body.get("plate") or "").strip().upper()
    if not plate:
        return jsonify({"error": "plate required"}), 400
    lat = body.get("lat")
    lon = body.get("lon")
    src = body.get("source") or "camera"
    raw = body.get("raw")
    s = PlateSighting(plate=plate, lat=(float(lat) if lat is not None else None),
                      lon=(float(lon) if lon is not None else None),
                      source=src, raw=(json.dumps(raw) if raw is not None else None))
    try:
        db.session.add(s)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "db_error"}), 500

    # check watchlist
    try:
        entry = Watchlist.query.filter(db.func.upper(Watchlist.plate) == plate).first()
        if entry:
            a = {
                "type": "watch_hit",
                "watch_id": entry.id,
                "watch_plate": entry.plate,
                "watch_label": entry.label,
                "plate": plate,
                "lat": s.lat,
                "lon": s.lon,
                "ts": s.ts.isoformat(),
                "source": s.source,
                "sighting_id": s.id
            }
            try:
                push_alert(a)
            except Exception:
                pass
            # emit to police room (real-time)
            try:
                socketio.emit('watch_hit', a, room='police')
            except Exception:
                pass
    except Exception:
        pass

    # emit the plate sighting to police UIs as well
    try:
        socketio.emit('plate_sighting', {"id": s.id, "plate": plate, "lat": s.lat, "lon": s.lon, "ts": s.ts.isoformat(), "source": s.source}, room='police')
    except Exception:
        pass

    return jsonify({"ok": True, "sighting_id": s.id})

# -------------------------
# WebSocket helpers for police clients
# -------------------------
@socketio.on('police_auth')
def police_auth(data):
    """Police UI should send: { 'token': 'ADMIN_TOKEN' } to join police room."""
    try:
        token = None
        if isinstance(data, dict):
            token = data.get('token') or ''
        # also allow session-based admin if they have cookie-based session
        if session.get('admin_logged'):
            join_room('police')
            emit('police_auth_ok', {"ok": True})
            return
        # check bearer/admin token
        if token and ADMIN_API_TOKEN and token == ADMIN_API_TOKEN:
            join_room('police')
            emit('police_auth_ok', {"ok": True})
            return
    except Exception:
        pass
    emit('police_auth_failed', {"ok": False})

@socketio.on('police_leave')
def police_leave(_data):
    try:
        leave_room('police')
    except Exception:
        pass

# -------------------------
# Minimal police dashboard (protected)
# -------------------------
POLICE_DASH_HTML = """
<!doctype html>
<html>
<head><meta charset="utf-8"><title>Police Dashboard</title>
<style>body{font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,Arial;margin:0;padding:0;background:linear-gradient(180deg,#07111f 0%, #0b1220 100%);color:#e5eefb;} #list{padding:12px;}</style>
</head>
<body>
  <h2 style="margin:12px 12px;">Police — Watch / Live Hits</h2>
  <div style="margin:12px;"><button id="btnAuth">Connect (admin token)</button></div>
  <div id="list"></div>
<script src="/socket.io/socket.io.js"></script>
<script>
  const s = io();
  const list = document.getElementById('list');
  document.getElementById('btnAuth').onclick = async () => {
    const token = prompt('Admin API token (or blank to use cookie session):');
    s.emit('police_auth', { token: token });
  };
  s.on('police_auth_ok', () => { list.innerHTML = '<div style="color:green;padding:8px;">Connected — listening for watch hits</div>'; });
  s.on('police_auth_failed', () => { alert('Auth failed'); });
  s.on('watch_hit', (d) => {
    const el = document.createElement('div');
    el.style.border = '1px solid #ccc'; el.style.padding='8px'; el.style.margin='8px';
    el.innerHTML = '<b>WATCH HIT</b> plate: '+(d.watch_plate||d.plate)+' <br/> label: '+(d.watch_label||'')+' <br/> at: '+(d.ts||'')+' <br/> loc: '+(d.lat||'')+','+(d.lon||'');
    list.prepend(el);
  });
  s.on('plate_sighting', (d) => {
    const el = document.createElement('div');
    el.style.border = '1px dashed #999'; el.style.padding='6px'; el.style.margin='8px';
    el.innerHTML = '<b>Plate sighting</b> '+d.plate+' @ '+(d.ts||'');
    list.prepend(el);
  });
</script>
</body>
</html>
"""


@app.route('/admin/users', methods=['GET', 'POST'])
@app.route('/admin/admins', methods=['GET', 'POST'])
def admin_admins():
    if _current_role() != 'admin':
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        password2 = request.form.get('password2') or ''
        role = (request.form.get('role') or 'admin').strip().lower()
        if role not in {'admin', 'police'}:
            role = 'admin'
        if not username or not password:
            flash("Username and password are required")
            return redirect(url_for('admin_admins'))
        if password != password2:
            flash("Passwords do not match")
            return redirect(url_for('admin_admins'))
        user, err = _create_account(username, password, role=role)
        if err:
            flash(err)
            return redirect(url_for('admin_admins'))
        flash(f"New {role} account added")
        return redirect(url_for('admin_admins'))

    admins = Admin.query.order_by(Admin.created_at.asc()).all()
    try:
        police_users = PoliceUser.query.order_by(PoliceUser.created_at.asc()).all()
    except Exception:
        police_users = []
    try:
        gk_users = GKUser.query.order_by(GKUser.created_at.asc()).all()
    except Exception:
        gk_users = []
    return _safe_render("""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>User Management</title>
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>
    body{font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,Arial;background:linear-gradient(180deg,#07111f 0%, #0b1220 100%);margin:0;padding:24px;color:#e5eefb;}
    .wrap{max-width:980px;margin:0 auto;}
    .card{background:#fff;border:1px solid #e2e8f0;border-radius:14px;padding:18px;box-shadow:0 10px 24px rgba(0,0,0,.05);margin-bottom:14px;}
    table{width:100%;border-collapse:collapse;}
    td,th{padding:10px;border-bottom:1px solid #e2e8f0;text-align:left;}
    input,select{width:100%;box-sizing:border-box;padding:12px;border:1px solid #e2e8f0;border-radius:10px;margin-top:8px;}
    button,a{display:inline-block;margin-top:14px;padding:12px 16px;border-radius:10px;border:0;background:#0b84ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer;}
    .muted{color:#64748b;font-size:14px;}
    .flash{background:#eff6ff;border:1px solid #dbeafe;color:#1e3a8a;padding:12px;border-radius:10px;margin-bottom:14px;}
    .top{display:flex;justify-content:space-between;gap:10px;align-items:center;flex-wrap:wrap;}
    .grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
    .full{grid-column:1/-1;}
    @media (max-width: 900px){ .grid{grid-template-columns:1fr;} }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Users & Access</h1>
      <div>
        <a href="{{ url_for('dashboard') }}">Dashboard</a>
        <a href="{{ url_for('all_vehicles') }}">All Vehicles</a>
        <a href="{{ url_for('admin_traffic') }}">Traffic Search</a>
        <a href="{{ url_for('admin_messages') }}">Messages</a>
        <a href="{{ url_for('admin_logout') }}">Logout</a>
      </div>
    </div>
    {% with messages = get_flashed_messages() %}
      {% if messages %}<div class="flash">{{ messages[0] }}</div>{% endif %}
    {% endwith %}
    <div class="card">
      <h2>Add admin, police or GK</h2>
      <form method="post" class="grid">
        <div class="full">
          <label>Role</label>
          <select name="role">
            <option value="admin">Admin</option>
            <option value="police">Police</option>
            <option value="gk">GK</option>
          </select>
        </div>
        <div>
          <label>Username</label>
          <input name="username" required>
        </div>
        <div>
          <label>Password</label>
          <input name="password" type="password" required>
        </div>
        <div class="full">
          <label>Confirm password</label>
          <input name="password2" type="password" required>
        </div>
        <div class="full">
          <button type="submit">Create account</button>
        </div>
      </form>
      <p class="muted">Registration stays closed to the public after the first admin. Only this page can add more accounts.</p>
    </div>
    <div class="grid">
      <div class="card">
        <h2>Admins</h2>
        <table>
          <thead><tr><th>Username</th><th>Created</th><th>Action</th></tr></thead>
          <tbody>
          {% for a in admins %}
            <tr>
              <td>{{ a.username }}</td>
              <td>{{ a.created_at.isoformat() if a.created_at else '' }}</td>
              <td>
                <form method="post" action="{{ url_for('admin_delete_user', role='admin', user_id=a.id) }}" style="margin:0">
                  <button type="submit" onclick="return confirm('Delete this admin?')">Delete</button>
                </form>
              </td>
            </tr>
          {% endfor %}
          </tbody>
        </table>
      </div>
      <div class="card">
        <h2>Police</h2>
        <table>
          <thead><tr><th>Username</th><th>Created</th><th>Action</th></tr></thead>
          <tbody>
          {% for p in police_users %}
            <tr>
              <td>{{ p.username }}</td>
              <td>{{ p.created_at.isoformat() if p.created_at else '' }}</td>
              <td>
                <form method="post" action="{{ url_for('admin_delete_user', role='police', user_id=p.id) }}" style="margin:0">
                  <button type="submit" onclick="return confirm('Delete this police user?')">Delete</button>
                </form>
              </td>
            </tr>
          {% endfor %}
          </tbody>
        </table>
      </div>
      <div class="card">
        <h2>GK</h2>
        <table>
          <thead><tr><th>Username</th><th>Created</th><th>Action</th></tr></thead>
          <tbody>
          {% for g in gk_users %}
            <tr>
              <td>{{ g.username }}</td>
              <td>{{ g.created_at.isoformat() if g.created_at else '' }}</td>
              <td>
                <form method="post" action="{{ url_for('admin_delete_user', role='gk', user_id=g.id) }}" style="margin:0">
                  <button type="submit" onclick="return confirm('Delete this GK user?')">Delete</button>
                </form>
              </td>
            </tr>
          {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</body>
</html>
""", admins=admins, police_users=police_users, gk_users=gk_users)



@app.route('/admin/users/delete/<role>/<int:user_id>', methods=['POST'])
def admin_delete_user(role, user_id):
    if _current_role() != 'admin':
        return redirect(url_for('admin_login'))

    role = (role or '').strip().lower()
    if role == 'admin':
        user = Admin.query.get_or_404(user_id)
        total_admins = Admin.query.count()
        if total_admins <= 1:
            flash("At least one admin must remain in the system.")
            return redirect(url_for('admin_admins'))
        # allow deleting the bootstrap admin too
        db.session.delete(user)
        db.session.commit()
        if session.get('username') == user.username and session.get('auth_role') == 'admin':
            session.clear()
            flash("That admin was removed. Please log in again.")
            return redirect(url_for('admin_login'))
        flash("Admin removed")
        return redirect(url_for('admin_admins'))

    if role == 'police':
        user = PoliceUser.query.get_or_404(user_id)
        db.session.delete(user)
        db.session.commit()
        flash("Police user removed")
        return redirect(url_for('admin_admins'))

    if role == 'gk':
        user = GKUser.query.get_or_404(user_id)
        db.session.delete(user)
        db.session.commit()
        flash("GK user removed")
        return redirect(url_for('admin_admins'))

    abort(400)

@app.route('/all-vehicles')
def all_vehicles():
    if _current_role() not in {'admin', 'police'}:
        return redirect(url_for('admin_login'))
    return _safe_render("""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>All Vehicles</title>
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
  <style>
    :root{
      --bg:#0b1220; --panel:#101a2d; --line:rgba(148,163,184,.16); --muted:#94a3b8; --text:#e5eefb;
      --accent:#8b5cf6; --accent2:#38bdf8; --good:#22c55e;
    }
    *{box-sizing:border-box}
    body{margin:0;font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,Arial;background:
      radial-gradient(circle at top left, rgba(139,92,246,.20), transparent 26%),
      radial-gradient(circle at top right, rgba(56,189,248,.18), transparent 24%),
      linear-gradient(180deg,#07111f 0%, #0b1220 100%);color:var(--text)}
    header{display:flex;gap:12px;align-items:center;flex-wrap:wrap;padding:18px 20px;background:rgba(16,26,45,.88);backdrop-filter: blur(14px);border-bottom:1px solid var(--line);position:sticky;top:0;z-index:20}
    header h1{margin:0;font-size:18px;letter-spacing:.2px}
    header a{color:#fff;text-decoration:none;font-weight:800;background:linear-gradient(135deg,rgba(139,92,246,.95),rgba(56,189,248,.95));padding:9px 14px;border-radius:999px;box-shadow:0 10px 24px rgba(56,189,248,.15)}
    header .hint{margin-left:auto;opacity:.9;font-size:13px;color:#dbeafe}
    .wrap{display:grid;grid-template-columns:390px 1fr;gap:16px;padding:16px;height:calc(100vh - 78px);box-sizing:border-box}
    .card{background:linear-gradient(180deg,rgba(16,26,45,.96),rgba(15,23,42,.96));border:1px solid var(--line);border-radius:24px;padding:16px;box-shadow:0 24px 64px rgba(0,0,0,.32);overflow:hidden}
    .stack{display:flex;flex-direction:column;height:100%;gap:14px}
    .searchbar{display:grid;grid-template-columns:1.1fr .7fr auto;gap:10px}
    .searchbar input,.searchbar select,.searchbar button{padding:12px 14px;border-radius:14px;border:1px solid var(--line);font-size:14px;background:#0b1324;color:var(--text)}
    .searchbar input::placeholder{color:#64748b}
    .searchbar button{background:linear-gradient(135deg,var(--accent),var(--accent2));color:#fff;border:none;cursor:pointer;font-weight:800}
    .statsGrid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}
    .statCard{border:1px solid var(--line);background:rgba(255,255,255,.03);border-radius:18px;padding:12px}
    .statCard .k{font-size:12px;color:var(--muted)}
    .statCard .v{font-size:22px;font-weight:800;margin-top:4px}
    #vehicleList{overflow:auto;flex:1;padding-right:4px}
    .vehicle{border:1px solid var(--line);border-radius:18px;padding:14px;margin-bottom:10px;cursor:pointer;background:rgba(255,255,255,.03);transition:transform .15s ease, border-color .15s ease, background .15s ease}
    .vehicle:hover{transform:translateY(-1px);border-color:rgba(56,189,248,.35)}
    .vehicle.selected{border-color:rgba(34,197,94,.7);background:rgba(34,197,94,.08)}
    .row{display:flex;justify-content:space-between;gap:10px;align-items:flex-start}
    .big{font-weight:800;font-size:15px}
    .muted{color:var(--muted);font-size:13px;line-height:1.5}
    .badge{display:inline-block;margin-top:8px;padding:5px 10px;border-radius:999px;background:rgba(56,189,248,.12);color:#7dd3fc;font-size:12px;font-weight:800;border:1px solid rgba(56,189,248,.24)}
    .badge.live{background:rgba(34,197,94,.15);color:#86efac;border-color:rgba(34,197,94,.25)}
    #map{height:100%;width:100%;border-radius:24px;overflow:hidden}
    .detail{margin-top:auto;padding-top:14px;border-top:1px solid var(--line)}
    .detailTop{display:flex;justify-content:space-between;gap:10px;align-items:center;flex-wrap:wrap}
    .detailTitle{font-size:20px;font-weight:800}
    .stats{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:12px}
    .stat{border:1px solid var(--line);border-radius:16px;padding:12px;background:rgba(255,255,255,.03)}
    .stat .k{color:var(--muted);font-size:12px}
    .stat .v{font-weight:800;margin-top:4px;word-break:break-word}
    .empty{padding:24px;border:1px dashed rgba(148,163,184,.35);border-radius:18px;text-align:center;color:var(--muted);margin-top:8px}
    @media (max-width: 980px){ .wrap{grid-template-columns:1fr;height:auto} #map{height:58vh} .searchbar{grid-template-columns:1fr} .statsGrid,.stats{grid-template-columns:1fr} }
  </style>
</head>
<body>
  <header>
    <h1>All Vehicles</h1>
    <a href="{{ url_for('dashboard') }}">Dashboard</a>
    <a href="{{ url_for('admin_traffic') }}">Traffic Search</a>
    <a href="{{ url_for('admin_admins') }}">Users</a>
    <a href="{{ url_for('admin_logout') }}">Logout</a>
    <span class="hint">Search by plate, owner, make, model, or status. Vehicles with no beacon still stay visible.</span>
  </header>

  <div class="wrap">
    <div class="card stack">
      <div>
        <div class="searchbar">
          <input id="q" placeholder="Search by plate, owner, make, model..." />
          <select id="field">
            <option value="all">All</option>
            <option value="plate">Plate</option>
            <option value="owner">Owner</option>
            <option value="model">Make / model</option>
            <option value="status">Status</option>
          </select>
          <button id="btnSearch">Search</button>
        </div>
        <div class="statsGrid" style="margin-top:12px;">
          <div class="statCard"><div class="k">Total</div><div class="v" id="statTotal">0</div></div>
          <div class="statCard"><div class="k">Live</div><div class="v" id="statLive">0</div></div>
          <div class="statCard"><div class="k">Offline</div><div class="v" id="statOffline">0</div></div>
          <div class="statCard"><div class="k">With beacon</div><div class="v" id="statBeacon">0</div></div>
        </div>
        <div class="muted" id="count" style="margin-top:10px;">Loading vehicles…</div>
      </div>

      <div id="vehicleList"></div>

      <div class="detail" id="detailBox">
        <div class="detailTop">
          <div>
            <div class="detailTitle" id="detailName">Select a vehicle</div>
            <div class="muted" id="detailMeta">Its last beacon and location will appear here.</div>
          </div>
          <span class="badge" id="detailStatus">—</span>
        </div>
        <div class="stats">
          <div class="stat"><div class="k">Plate</div><div class="v" id="detailPlate">—</div></div>
          <div class="stat"><div class="k">Owner</div><div class="v" id="detailOwner">—</div></div>
          <div class="stat"><div class="k">Speed</div><div class="v" id="detailSpeed">—</div></div>
        </div>
      </div>
    </div>

    <div class="card"><div id="map"></div></div>
  </div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script>
  const map = L.map('map', { center:[0,0], zoom:2, preferCanvas:true });
  L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', { maxZoom:19 }).addTo(map);
  let vehicles = [];
  let selected = null;
  let marker = null;
  let circle = null;

  const qEl = document.getElementById('q');
  const fieldEl = document.getElementById('field');
  const listEl = document.getElementById('vehicleList');
  const countEl = document.getElementById('count');
  const statTotal = document.getElementById('statTotal');
  const statLive = document.getElementById('statLive');
  const statOffline = document.getElementById('statOffline');
  const statBeacon = document.getElementById('statBeacon');

  function escapeHtml(s) {
    if (!s) return '';
    return String(s).replace(/[&<>"'`]/g, (c) => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;','`':'&#96;'})[c]);
  }

  async function loadVehicles() {
    const q = qEl.value.trim();
    const field = fieldEl.value;
    const res = await fetch(`/admin/vehicles?q=${encodeURIComponent(q)}&field=${encodeURIComponent(field)}`, { cache: 'no-store' });
    if (!res.ok) {
      countEl.innerText = 'Login required';
      return;
    }
    const data = await res.json();
    vehicles = data.vehicles || [];
    const total = data.count ?? vehicles.length;
    const live = data.live ?? vehicles.filter(v => v.online).length;
    const offline = Math.max(0, total - live);
    const withBeacon = vehicles.filter(v => !!(v.last_snapshot && v.last_snapshot.ts)).length;
    countEl.innerText = `${total} vehicle(s) shown`;
    statTotal.innerText = total;
    statLive.innerText = live;
    statOffline.innerText = offline;
    statBeacon.innerText = withBeacon;
    renderVehicles();
    if (vehicles.length && !selected) selectVehicle(vehicles[0].id);
    if (selected && !vehicles.some(v => v.id === selected)) selected = null;
  }

  function vehicleLabel(v) {
    return v.label || [v.make, v.model].filter(Boolean).join(' ').trim() || v.car_name || v.car_model || v.id;
  }

  function renderVehicles() {
    listEl.innerHTML = '';
    if (!vehicles.length) {
      listEl.innerHTML = '<div class="empty">No vehicles match the current search.</div>';
      return;
    }
    for (const v of vehicles) {
      const div = document.createElement('div');
      div.className = 'vehicle' + (v.id === selected ? ' selected' : '');
      const last = v.last_snapshot || {};
      const when = last.ts ? new Date(last.ts).toLocaleString() : 'no beacon yet';
      const spd = (last.speed_mps != null) ? ((last.speed_mps * 3.6).toFixed(1) + ' km/h') : '—';
      const status = v.online ? 'live' : 'offline';
      div.innerHTML = `
        <div class="row">
          <div style="min-width:0;">
            <div class="big">${escapeHtml(vehicleLabel(v))}</div>
            <div class="muted">${escapeHtml(v.owner || 'No owner recorded')}</div>
          </div>
          <div style="text-align:right;">
            <div class="muted">${escapeHtml(v.plate || '—')}</div>
            <div class="badge ${v.online ? 'live' : ''}">${escapeHtml(status)}</div>
          </div>
        </div>
        <div class="muted" style="margin-top:8px;">Beacon: ${escapeHtml(when)} · ${escapeHtml(spd)} · source: ${escapeHtml(v.source || 'none')}</div>
      `;
      div.addEventListener('click', () => selectVehicle(v.id));
      listEl.appendChild(div);
    }
  }

  function selectVehicle(id) {
    selected = id;
    renderVehicles();
    const v = vehicles.find(x => x.id === id);
    if (!v) return;
    const last = v.last_snapshot || {};
    document.getElementById('detailName').innerText = vehicleLabel(v);
    document.getElementById('detailMeta').innerText = last.ts ? new Date(last.ts).toLocaleString() : 'No beacon yet';
    document.getElementById('detailPlate').innerText = v.plate || '—';
    document.getElementById('detailOwner').innerText = v.owner || '—';
    document.getElementById('detailSpeed').innerText = last.speed_mps != null ? ((last.speed_mps * 3.6).toFixed(1) + ' km/h') : '—';
    document.getElementById('detailStatus').innerText = v.online ? 'LIVE' : 'OFFLINE';

    if (typeof last.lat === 'number' && typeof last.lon === 'number') {
      if (!marker) marker = L.marker([last.lat, last.lon]).addTo(map);
      else marker.setLatLng([last.lat, last.lon]);
      if (!circle) circle = L.circle([last.lat, last.lon], { radius: 25 }).addTo(map);
      else circle.setLatLng([last.lat, last.lon]);
      marker.bindPopup(`<strong>${escapeHtml(vehicleLabel(v))}</strong><br/>${escapeHtml(v.plate || '')}`).openPopup();
      map.setView([last.lat, last.lon], 15, { animate:true });
    }
  }

  document.getElementById('btnSearch').addEventListener('click', loadVehicles);
  qEl.addEventListener('keydown', (e) => { if (e.key === 'Enter') loadVehicles(); });
  qEl.addEventListener('input', () => {
    clearTimeout(window._vehTimer);
    window._vehTimer = setTimeout(loadVehicles, 200);
  });
  fieldEl.addEventListener('change', loadVehicles);

  loadVehicles();
</script>
</body>
</html>
""")
@app.route('/admin/traffic', methods=['GET', 'POST'])
def admin_traffic():
    if _current_role() != 'admin':
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        scope = (request.form.get('scope') or 'custom').strip().lower()
        notes = (request.form.get('notes') or '').strip()
        zone_id = (request.form.get('zone_id') or '').strip()
        zone = TrafficZone.query.filter_by(id=zone_id).first() if zone_id else TrafficZone()

        if name:
            zone.name = name
        zone.scope = scope if scope in {'county', 'constituency', 'region', 'road', 'national', 'custom'} else 'custom'
        zone.notes = notes or None

        for key in ('center_lat', 'center_lon', 'radius_m'):
            val = (request.form.get(key) or '').strip()
            if val:
                try:
                    setattr(zone, key, float(val))
                except Exception:
                    pass

        if not zone.name:
            zone.name = zone.scope.title() if zone.scope else 'Traffic Zone'

        db.session.add(zone)
        db.session.commit()
        flash('Traffic zone saved')
        return redirect(url_for('admin_traffic', zone_id=zone.id))

    zone_id = (request.args.get('zone_id') or '').strip()
    road_id = (request.args.get('road_id') or '').strip()
    zone = TrafficZone.query.filter_by(id=zone_id).first() if zone_id else None
    if not zone and (request.args.get('scope') or '').lower() == 'national':
        zone = TrafficZone(name='National Traffic', scope='national')

    selected_road = Road.query.filter_by(id=road_id).first() if road_id else None
    if not zone and selected_road:
        zone = TrafficZone(
            name=selected_road.name,
            scope='road',
            center_lat=selected_road.center_lat,
            center_lon=selected_road.center_lon,
            radius_m=selected_road.radius_m,
            notes=f'Road preview: {selected_road.name}'
        )

    vehicles, counts = _traffic_snapshot(zone)
    zones = TrafficZone.query.order_by(TrafficZone.created_at.desc()).all()
    roads = Road.query.order_by(Road.created_at.desc()).all()
    return render_template_string("""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Traffic Search</title>
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
  <style>
    :root{--bg:#08111f;--card:#0f172a;--line:rgba(148,163,184,.15);--muted:#94a3b8;--text:#e6eefc;--accent:#8b5cf6;--accent2:#38bdf8;--accent3:#22c55e;}
    *{box-sizing:border-box}
    body{font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,Arial;margin:0;background:radial-gradient(circle at top left, rgba(139,92,246,.22), transparent 28%),radial-gradient(circle at top right, rgba(56,189,248,.18), transparent 26%),linear-gradient(180deg,#050b14 0%, #08111f 100%);color:var(--text)}
    .wrap{max-width:1340px;margin:0 auto;padding:18px}
    .hero{display:flex;justify-content:space-between;gap:12px;align-items:center;flex-wrap:wrap;background:linear-gradient(135deg,rgba(139,92,246,.16),rgba(56,189,248,.12));border:1px solid var(--line);border-radius:26px;padding:18px 20px;box-shadow:0 24px 64px rgba(0,0,0,.28)}
    .hero h1{margin:0;font-size:28px}
    .hero .muted{color:var(--muted);margin-top:6px;line-height:1.5}
    .hero .actions a{display:inline-block;margin-left:8px;margin-top:8px;padding:10px 14px;border-radius:999px;background:rgba(255,255,255,.08);color:#fff;text-decoration:none;font-weight:800;border:1px solid var(--line)}
    .grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:16px}
    .card{background:linear-gradient(180deg,rgba(16,28,49,.96),rgba(15,23,42,.96));border:1px solid var(--line);border-radius:24px;padding:16px;box-shadow:0 22px 50px rgba(0,0,0,.28)}
    .full{grid-column:1/-1}
    .sectionTitle{margin:0 0 12px;font-size:18px;font-weight:800}
    .muted{color:var(--muted);font-size:13px;line-height:1.5}
    input,select,textarea,button,a.linkBtn{width:100%;box-sizing:border-box;padding:12px 14px;border:1px solid var(--line);border-radius:14px;margin-top:8px;background:#0b1324;color:var(--text)}
    button,a.linkBtn{display:inline-flex;justify-content:center;align-items:center;text-decoration:none;font-weight:800;cursor:pointer;border:none;background:linear-gradient(135deg,var(--accent),var(--accent2));color:#fff}
    .pill{display:inline-flex;align-items:center;gap:8px;padding:8px 12px;border-radius:999px;background:rgba(56,189,248,.12);color:#d7f3ff;font-weight:800;margin:6px 6px 0 0;border:1px solid rgba(56,189,248,.18)}
    .roadCard{border:1px solid var(--line);border-radius:18px;padding:14px;background:rgba(255,255,255,.03);margin-top:10px}
    .roadTop{display:flex;justify-content:space-between;gap:10px;align-items:flex-start}
    .roadName{font-weight:800;font-size:15px}
    .roadMeta{font-size:12px;color:var(--muted);margin-top:4px;line-height:1.45}
    .roadActions{display:flex;gap:8px;flex-wrap:wrap;margin-top:10px}
    .roadActions a,.roadActions button{width:auto;flex:1;min-width:120px;padding:10px 12px;border-radius:12px}
    table{width:100%;border-collapse:collapse;margin-top:8px;background:rgba(255,255,255,.02);border-radius:18px;overflow:hidden}
    td,th{padding:11px 10px;border-bottom:1px solid rgba(148,163,184,.12);text-align:left;font-size:13px;vertical-align:top}
    th{font-size:12px;text-transform:uppercase;letter-spacing:.04em;color:#c7d2fe}
    .tableWrap{overflow:auto;border:1px solid var(--line);border-radius:18px}
    .summaryGrid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}
    .summaryCard{border:1px solid var(--line);border-radius:18px;padding:12px;background:rgba(255,255,255,.03)}
    .summaryCard .k{font-size:12px;color:var(--muted)}
    .summaryCard .v{font-size:22px;font-weight:800;margin-top:4px}
    @media (max-width: 980px){ .grid{grid-template-columns:1fr} .summaryGrid{grid-template-columns:1fr} }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="hero">
      <div>
        <h1>Traffic Search</h1>
        <div class="muted">Search by zone, road, county, constituency, region, or national view. Created roads now appear directly under saved zones for quick operations.</div>
      </div>
      <div class="actions">
        <a href="{{ url_for('dashboard') }}">Dashboard</a>
        <a href="{{ url_for('all_vehicles') }}">All Vehicles</a>
        <a href="{{ url_for('admin_admins') }}">Users</a>
        <a href="{{ url_for('admin_logout') }}">Logout</a>
      </div>
    </div>

    {% with messages = get_flashed_messages() %}
      {% if messages %}<div class="card" style="margin-top:16px;border-color:rgba(34,197,94,.28);background:rgba(34,197,94,.06);">{{ messages[0] }}</div>{% endif %}
    {% endwith %}

    <div class="grid">
      <div class="card">
        <h2 class="sectionTitle">Search zone</h2>
        <form method="get">
          <label>Saved zone</label>
          <select name="zone_id">
            <option value="">-- pick a saved zone --</option>
            {% for z in zones %}
              <option value="{{ z.id }}" {% if zone and z.id == zone.id %}selected{% endif %}>{{ z.name }} ({{ z.scope }})</option>
            {% endfor %}
          </select>
          <label>Saved road</label>
          <select name="road_id">
            <option value="">-- pick a created road --</option>
            {% for r in roads %}
              <option value="{{ r.id }}" {% if selected_road and r.id == selected_road.id %}selected{% endif %}>{{ r.name }} · {{ '%.1f km/h'|format(r.speed_limit_kmh) }}</option>
            {% endfor %}
          </select>
          <label>National view</label>
          <select name="scope">
            <option value="">-- keep selected zone --</option>
            <option value="national">National</option>
          </select>
          <button type="submit">Search traffic</button>
        </form>

        <form method="post" style="margin-top:18px;">
          <h3 style="margin:4px 0 10px;">Add / update zone</h3>
          <input type="hidden" name="zone_id" value="{{ zone.id if zone else '' }}">
          <label>Name</label>
          <input name="name" value="{{ zone.name if zone else '' }}" placeholder="e.g. Nairobi CBD">
          <label>Scope</label>
          <select name="scope">
            {% set scopes = ['county','constituency','region','road','national','custom'] %}
            {% for s in scopes %}
              <option value="{{ s }}" {% if zone and zone.scope == s %}selected{% endif %}>{{ s }}</option>
            {% endfor %}
          </select>
          <label>Center latitude</label>
          <input name="center_lat" value="{{ zone.center_lat if zone and zone.center_lat is not none else '' }}">
          <label>Center longitude</label>
          <input name="center_lon" value="{{ zone.center_lon if zone and zone.center_lon is not none else '' }}">
          <label>Radius metres</label>
          <input name="radius_m" value="{{ zone.radius_m if zone and zone.radius_m is not none else '' }}">
          <label>Notes</label>
          <textarea name="notes" rows="3">{{ zone.notes if zone and zone.notes else '' }}</textarea>
          <button type="submit">Save zone</button>
        </form>
      </div>

      <div class="card">
        <h2 class="sectionTitle">Created roads</h2>
        <div class="muted">These are the roads you created in the road manager. Use them immediately for traffic search or convert them into a zone.</div>
        <div style="margin-top:10px;">
          {% for r in roads %}
            <div class="roadCard">
              <div class="roadTop">
                <div>
                  <div class="roadName">{{ r.name }}</div>
                  <div class="roadMeta">Speed limit: {{ '%.1f km/h'|format(r.speed_limit_kmh) }} · Radius: {{ '%.0f m'|format(r.radius_m or 0) }} · {{ r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '' }}</div>
                  <div class="roadMeta">Center: {{ '%.5f'|format(r.center_lat) if r.center_lat is not none else '—' }}, {{ '%.5f'|format(r.center_lon) if r.center_lon is not none else '—' }}</div>
                </div>
                <span class="pill">road</span>
              </div>
              <div class="roadActions">
                <a class="linkBtn" href="{{ url_for('admin_traffic', road_id=r.id) }}">Use for search</a>
                <button type="button" class="smallbtn" onclick="fillZoneFromRoad('{{ r.name|e }}', '{{ r.center_lat if r.center_lat is not none else '' }}', '{{ r.center_lon if r.center_lon is not none else '' }}', '{{ r.radius_m if r.radius_m is not none else '' }}')">Fill zone form</button>
              </div>
            </div>
          {% endfor %}
          {% if not roads %}
            <div class="roadCard">No roads created yet.</div>
          {% endif %}
        </div>
      </div>

      <div class="card full">
        <div class="summaryGrid">
          <div class="summaryCard"><div class="k">Total vehicles</div><div class="v">{{ vehicles|length }}</div></div>
          <div class="summaryCard"><div class="k">Cars</div><div class="v">{{ counts.get('car', 0) }}</div></div>
          <div class="summaryCard"><div class="k">Trucks</div><div class="v">{{ counts.get('truck', 0) }}</div></div>
          <div class="summaryCard"><div class="k">Buses</div><div class="v">{{ counts.get('bus', 0) }}</div></div>
          <div class="summaryCard"><div class="k">Vans</div><div class="v">{{ counts.get('van', 0) }}</div></div>
          <div class="summaryCard"><div class="k">Motorcycles</div><div class="v">{{ counts.get('motorcycle', 0) }}</div></div>
        </div>
      </div>

      <div class="card full">
        <h2 class="sectionTitle">Vehicles in this traffic area</h2>
        <div class="tableWrap">
          <table>
            <thead><tr><th>Plate</th><th>Vehicle</th><th>Owner</th><th>Type</th><th>Speed</th><th>Last position</th></tr></thead>
            <tbody>
              {% for v in vehicles %}
                <tr>
                  <td>{{ v.plate or '—' }}</td>
                  <td>{{ v.label or v.car_name or v.car_model or v.device_id[:8] }}</td>
                  <td>{{ v.owner or '—' }}</td>
                  <td>{{ v.vehicle_type }}</td>
                  <td>{{ '%.1f km/h'|format(v.last_snapshot.speed_mps * 3.6) if v.last_snapshot and v.last_snapshot.speed_mps is not none else '—' }}</td>
                  <td>{{ v.last_snapshot.lat if v.last_snapshot and v.last_snapshot.lat is not none else '—' }}, {{ v.last_snapshot.lon if v.last_snapshot and v.last_snapshot.lon is not none else '—' }}</td>
                </tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
      </div>
    </div>
  </div>
  <script>
    function fillZoneFromRoad(name, lat, lon, radius) {
      const nameEl = document.querySelector('input[name="name"]');
      const latEl = document.querySelector('input[name="center_lat"]');
      const lonEl = document.querySelector('input[name="center_lon"]');
      const radiusEl = document.querySelector('input[name="radius_m"]');
      const scopeEl = document.querySelector('select[name="scope"]');
      if (nameEl) nameEl.value = name || '';
      if (latEl) latEl.value = lat || '';
      if (lonEl) lonEl.value = lon || '';
      if (radiusEl) radiusEl.value = radius || '';
      if (scopeEl) scopeEl.value = 'road';
      window.scrollTo({ top: 0, behavior: 'smooth' });
    }
  </script>
</body>
</html>
""", zone=zone, zones=zones, vehicles=vehicles, counts=counts, roads=roads, selected_road=selected_road)
@app.route('/admin/vehicles')
def admin_vehicles():
    if _current_role() not in {'admin', 'police'}:
        return jsonify({"error": "Login required"}), 401

    q = (request.args.get("q") or "").strip().lower()
    field = (request.args.get("field") or "all").strip().lower()
    devices = Device.query.all()
    out = []

    for d in devices:
        hay_plate = (d.plate or "").lower()
        hay_owner = (d.owner or "").lower()
        hay_name = ((d.car_name or "") + " " + (d.car_model or "")).lower()

        match = True
        if q:
            if field == "plate":
                match = q in hay_plate
            elif field == "owner":
                match = q in hay_owner
            else:
                match = (q in hay_plate) or (q in hay_owner) or (q in hay_name)

        if not match:
            continue

        snap = Snapshot.query.filter_by(device_id=d.id).order_by(Snapshot.ts.desc()).first()
        last = None
        if snap:
            last = {
                "ts": snap.ts.isoformat() if snap.ts else None,
                "lat": snap.lat,
                "lon": snap.lon,
                "speed_mps": round(snap.speed_mps or 0.0, 3),
                "bearing_deg": round(snap.bearing_deg or 0.0, 1),
            }

        out.append({
            "id": d.id,
            "owner": d.owner,
            "car_name": d.car_name,
            "car_model": d.car_model,
            "plate": d.plate,
            "connected": bool(connected_sockets.get(d.id)),
            "last_snapshot": last,
        })

    out.sort(key=lambda item: item.get("last_snapshot", {}).get("ts") or "", reverse=True)
    return jsonify({"vehicles": out, "count": len(out)})

@app.route('/police')
def police_dashboard():
    if _current_role() not in {'admin', 'police'}:
        return redirect(url_for('admin_login'))
    return _safe_render(POLICE_DASH_HTML)
# -------------------------
# End of police/watch block
# -------------------------

# -------------------------
# Admin endpoints for alerts & jams (requires admin session or ADMIN_API_TOKEN)
# -------------------------
@app.route('/admin/alerts', methods=['GET'])
def admin_list_alerts():
    require_admin_api()
    limit = int(request.args.get("limit", 100))
    return jsonify({"alerts": list_alerts(limit=limit)})

@app.route('/admin/alerts/clear', methods=['POST'])
def admin_clear_alerts():
    require_admin_api()
    clear_alerts()
    return jsonify({"ok": True})

@app.route('/admin/jams', methods=['GET'])
def admin_list_jams():
    require_admin_api()
    limit = int(request.args.get("limit", 100))
    return jsonify({"jams": list_jams(limit=limit)})

@app.route('/admin/jams/clear', methods=['POST'])
def admin_clear_jams():
    require_admin_api()
    clear_jams()
    return jsonify({"ok": True})

# -------------------------
# Socket.IO events (preserved + auto-restore on register)
# -------------------------
@socketio.on('connect')
def ws_connect():
    sid = request.sid
    app.logger.debug('Socket connected: %s', sid)
    emit('connected', {'sid': sid})

@socketio.on('register')
def ws_register(data):
    sid = request.sid
    device_id = data.get('device_id')
    token = data.get('token')
    if not device_id or not token:
        emit('error', {'error': 'device_id and token required'})
        return
    device = Device.query.filter_by(id=device_id, token=token, revoked=False).first()
    if not device:
        # Attempt to restore silently (device kept token)
        device = restore_device_if_missing(device_id, token, payload=data)
        if not device:
            emit('error', {'error': 'invalid token/device'})
            return
    sids = connected_sockets.get(device_id)
    if not sids:
        connected_sockets[device_id] = set()
    connected_sockets[device_id].add(sid)
    join_room(sid)
    emit('registered', {'ok': True, 'device_id': device_id})
    app.logger.info('Device %s registered socket %s', device_id, sid)

@socketio.on('disconnect')
def ws_disconnect():
    sid = request.sid
    app.logger.debug('Socket disconnected: %s', sid)
    to_remove = []
    for dev, sids in list(connected_sockets.items()):
        if sid in sids:
            sids.discard(sid)
        if not sids:
            to_remove.append(dev)
    for d in to_remove:
        connected_sockets.pop(d, None)

@socketio.on('get_nearby')
def ws_get_nearby(data):
    device_id = data.get('device_id')
    token = data.get('token')
    device = Device.query.filter_by(id=device_id, token=token, revoked=False).first()
    if not device:
        # try restore (best-effort)
        device = restore_device_if_missing(device_id, token, payload=data)
    if not device:
        emit('error', {'error': 'invalid device/token'})
        return
    try:
        payload = compute_nearby_for_device(device_id, radius_m=NEARBY_DEFAULT_RADIUS_M)
        emit('nearby', payload)
    except Exception as e:
        emit('error', {'error': str(e)})

# -------------------------
# Jam detection background worker (new)
# -------------------------
def jam_detector_once():
    """
    Simple clustering: gather recent active snapshots and find clusters of devices
    within JAM_CLUSTER_RADIUS_M whose average speed is below JAM_SPEED_THRESHOLD_MPS.
    Produces a jam dict: { id, ts, lat, lon, count, avg_speed, device_ids }
    """
    try:
        cutoff = datetime.utcnow() - timedelta(seconds=CLEANUP_STALE_SECONDS)
        snaps = Snapshot.query.filter(Snapshot.ts >= cutoff).all()
        # add in-memory cache snapshot fallback for devices not in DB recent snaps
        with active_devices_lock:
            for did, entry in active_devices.items():
                if entry.get("ts") and entry.get("ts") >= cutoff:
                    # skip if DB already has a recent snapshot for same device
                    if any(s.device_id == did for s in snaps):
                        continue
                    class _T:
                        pass
                    t = _T()
                    t.device_id = did
                    t.lat = entry.get("lat")
                    t.lon = entry.get("lon")
                    t.speed_mps = entry.get("speed_mps") or 0.0
                    t.bearing_deg = entry.get("bearing_deg") or 0.0
                    t.ts = entry.get("ts")
                    snaps.append(t)
        if not snaps:
            return

        used = set()
        clusters = []
        for s in snaps:
            if s.device_id in used:
                continue
            # build cluster around s
            members = [s]
            for t in snaps:
                if t.device_id == s.device_id or t.device_id in used:
                    continue
                d = haversine_m(s.lat, s.lon, t.lat, t.lon)
                if d <= JAM_CLUSTER_RADIUS_M:
                    members.append(t)
            for m in members:
                used.add(m.device_id)
            # evaluate cluster
            count = len(members)
            avg_speed = sum([(m.speed_mps or 0.0) for m in members]) / max(1, count)
            if count >= JAM_MIN_DEVICES and avg_speed <= JAM_SPEED_THRESHOLD_MPS:
                # compute centroid
                avg_lat = sum([m.lat for m in members]) / count
                avg_lon = sum([m.lon for m in members]) / count
                cluster = {
                    "id": uuid.uuid4().hex,
                    "ts": datetime.utcnow().isoformat(),
                    "lat": round(avg_lat, 6),
                    "lon": round(avg_lon, 6),
                    "count": count,
                    "avg_speed_mps": round(avg_speed, 2),
                    "device_ids": [m.device_id for m in members]
                }
                clusters.append(cluster)

        # push clusters to jams_store and optionally emit to involved device sockets
        for jam in clusters:
            push_jam(jam)
            # emit to each device in jam that has sockets
            for did in jam["device_ids"]:
                try:
                    send_ws_to_device(did, "jam_alert", jam)
                except Exception:
                    pass
    except Exception:
        app.logger.exception("jam_detector_once error")

def jam_detector_loop():
    while True:
        try:
            jam_detector_once()
        except Exception:
            app.logger.exception("jam_detector_loop error")
        time.sleep(JAM_DETECT_INTERVAL_S)

init_db()

# -------------------------
# CLI entry & startup hooks
# -------------------------
if __name__ == "__main__":
    init_db()
    # start jam detector background thread
    try:
        t = threading.Thread(target=jam_detector_loop, daemon=True)
        t.start()
        app.logger.info("Jam detector thread started.")
    except Exception:
        app.logger.exception("Failed to start jam detector thread.")

    socketio.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=os.environ.get("FLASK_DEBUG", "0") == "1")
# End of file
